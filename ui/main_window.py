import tkinter as tk
from tkinter import ttk, messagebox
from config import COLORS, FONTS, WINDOW_GEOMETRY, WINDOW_BG
from .products_page import ProductsPage
from .meal_registration_page import MealRegistrationPage
from PIL import Image, ImageTk
from datetime import datetime
import os

class MainWindow:
    def __init__(self, root, data_manager):
        self.root = root
        self.data_manager = data_manager
        
        self.root.title("Quản Lý Thực Phẩm")
        self.root.geometry(WINDOW_GEOMETRY)
        self.root.configure(bg=WINDOW_BG)
        
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        self.root.minsize(800, 600)
        self.root.state('zoomed')
        
        try:
            icon_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "images", "app_icon.png")
            icon = Image.open(icon_path)
            icon_photo = ImageTk.PhotoImage(icon)
            self.root.iconphoto(True, icon_photo)
        except Exception as e:
            print(f"Could not load icon: {e}")
        
        self.show_main_menu()
    
    def lighten_color(self, hex_color, factor=0.2):
        hex_color = hex_color.lstrip('#')
        r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        r = int(r + (255 - r) * factor)
        g = int(g + (255 - g) * factor)
        b = int(b + (255 - b) * factor)
        return f'#{r:02x}{g:02x}{b:02x}'
    
    def bind_hover(self, button, original_color):
        hover_color = self.lighten_color(original_color)
        button.bind('<Enter>', lambda e: button.config(bg=hover_color))
        button.bind('<Leave>', lambda e: button.config(bg=original_color))
    
    def create_gradient(self, width, height, *colors):
        gradient = Image.new('RGB', (width, height))
        draw = gradient.load()
        
        rgb_colors = []
        for color in colors:
            color = color.lstrip('#')
            rgb = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
            rgb_colors.append(rgb)
        
        num_colors = len(rgb_colors)
        for y in range(height):
            ratio = y / height * (num_colors - 1)
            index = int(ratio)
            local_ratio = ratio - index
            
            if index >= num_colors - 1:
                r, g, b = rgb_colors[-1]
            else:
                r1, g1, b1 = rgb_colors[index]
                r2, g2, b2 = rgb_colors[index + 1]
                r = int(r1 + (r2 - r1) * local_ratio)
                g = int(g1 + (g2 - g1) * local_ratio)
                b = int(b1 + (b2 - b1) * local_ratio)
            
            for x in range(width):
                draw[x, y] = (r, g, b)
        
        return gradient
    
    def show_main_menu(self):
        for widget in self.root.winfo_children():
            widget.destroy()
        
        canvas = tk.Canvas(self.root, highlightthickness=0)
        canvas.pack(expand=True, fill='both')
        
        resize_timer = None
        last_size = [0, 0]
        
        def resize_canvas(event=None):
            width = self.root.winfo_width()
            height = self.root.winfo_height()
            
            if width == last_size[0] and height == last_size[1]:
                return
            
            last_size[0] = width
            last_size[1] = height
            
            canvas.delete('all')
            
            gradient_img = self.create_gradient(width, height, '#106e49', '#33a951', '#9bcc70')
            gradient_photo = ImageTk.PhotoImage(gradient_img)
            canvas.create_image(0, 0, anchor='nw', image=gradient_photo)
            canvas.image = gradient_photo
            
            try:
                icon_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "images", "app_icon.png")
                icon_img = Image.open(icon_path).convert("RGBA")
                
                icon_size = int(min(width, height) * 0.95)
                icon_img = icon_img.resize((icon_size, icon_size), Image.Resampling.LANCZOS)
                
                alpha = icon_img.split()[3]
                alpha = alpha.point(lambda p: int(p * 0.1))
                icon_img.putalpha(alpha)
                
                icon_photo = ImageTk.PhotoImage(icon_img)
                canvas.create_image(width // 2, height // 2, image=icon_photo)
                canvas.icon_image = icon_photo

            except Exception as e:
                print(f"Could not load background icon: {e}")
            
            self.draw_menu_content(canvas, width, height)
        
        def on_resize(event=None):
            nonlocal resize_timer
            if resize_timer:
                self.root.after_cancel(resize_timer)
            resize_timer = self.root.after(50, resize_canvas)
        
        self.root.after(100, resize_canvas)
        self.root.bind('<Configure>', on_resize)
    
    def draw_menu_content(self, canvas, width, height):
        center_x = width // 2
        center_y = height // 2
        
        canvas.create_text(
            center_x, 80,
            text="ỨNG DỤNG QUẢN LÝ THỰC PHẨM CP BVTV PHÚ NÔNG",
            font=FONTS['title'],
            fill='white'
        )
        
        btn_products_frame = tk.Frame(canvas, bg='white', bd=2)
        btn_products = tk.Button(
            btn_products_frame,
            text="CÁC MẶT HÀNG NÔNG SẢN - THỰC PHẨM",
            font=FONTS['button_large'],
            bg=COLORS['primary'],
            fg=COLORS['white'],
            width=35,
            height=2,
            command=self.show_products_page,
            cursor='hand2',
            bd=0,
            relief='flat'
        )
        btn_products.pack()
        self.bind_hover(btn_products, COLORS['primary'])
        canvas.create_window(center_x, center_y - 170, window=btn_products_frame)
        
        btn_meal_registration_frame = tk.Frame(canvas, bg='white', bd=2)
        btn_meal_registration = tk.Button(
            btn_meal_registration_frame,
            text="DANH SÁCH CNV ĐĂNG KÝ SUẤT ĂN",
            font=FONTS['button_large'],
            bg=COLORS['success'],
            fg=COLORS['white'],
            width=35,
            height=2,
            command=self.show_meal_registration_page,
            cursor='hand2',
            bd=0,
            relief='flat'
        )
        btn_meal_registration.pack()
        self.bind_hover(btn_meal_registration, COLORS['success'])
        canvas.create_window(center_x, center_y - 90, window=btn_meal_registration_frame)
        
        btn_update_employee_frame = tk.Frame(canvas, bg='white', bd=2)
        btn_update_employee = tk.Button(
            btn_update_employee_frame,
            text="CẬP NHẬT DANH SÁCH NHÂN VIÊN",
            font=FONTS['button_large'],
            bg=COLORS['pink'],
            fg=COLORS['white'],
            width=35,
            height=2,
            command=self.show_employee_management,
            cursor='hand2',
            bd=0,
            relief='flat'
        )
        btn_update_employee.pack()
        self.bind_hover(btn_update_employee, COLORS['pink'])
        canvas.create_window(center_x, center_y - 10, window=btn_update_employee_frame)
        
        btn_menu_management_frame = tk.Frame(canvas, bg='white', bd=2)
        btn_menu_management = tk.Button(
            btn_menu_management_frame,
            text="MENU",
            font=FONTS['button_large'],
            bg=COLORS['warning'],
            fg=COLORS['white'],
            width=35,
            height=2,
            command=self.show_menu_management,
            cursor='hand2',
            bd=0,
            relief='flat'
        )
        btn_menu_management.pack()
        self.bind_hover(btn_menu_management, COLORS['warning'])
        canvas.create_window(center_x, center_y + 70, window=btn_menu_management_frame)
        btn_order_frame = tk.Frame(canvas, bg='white', bd=2)
        btn_order = tk.Button(
            btn_order_frame,
            text="ĐẶT HÀNG",
            font=FONTS['button_large'],
            bg='#ff6b6b',
            fg=COLORS['white'],
            width=35,
            height=2,
            command=self.show_order_management,
            cursor='hand2',
            bd=0,
            relief='flat'
        )
        btn_order.pack()
        self.bind_hover(btn_order, '#ff6b6b')
        canvas.create_window(center_x, center_y + 150, window=btn_order_frame)
        
        btn_statistics_frame = tk.Frame(canvas, bg='white', bd=2)
        btn_statistics = tk.Button(
            btn_statistics_frame,
            text="THỐNG KÊ",
            font=FONTS['button_large'],
            bg=COLORS['info'],
            fg=COLORS['white'],
            width=35,
            height=2,
            command=self.show_statistics,
            cursor='hand2',
            bd=0,
            relief='flat'
        )
        btn_statistics.pack()
        self.bind_hover(btn_statistics, COLORS['info'])
        canvas.create_window(center_x, center_y + 230, window=btn_statistics_frame)
        
        exit_btn_frame = tk.Frame(canvas, bg='white', bd=2)
        exit_btn = tk.Button(
            exit_btn_frame,
            text="Thoát",
            font=FONTS['section'],
            bg=COLORS['danger'],
            fg=COLORS['white'],
            width=15,
            command=self.root.quit,
            bd=0,
            relief='flat'
        )
        exit_btn.pack()
        self.bind_hover(exit_btn, COLORS['danger'])
        canvas.create_window(center_x, center_y + 310, window=exit_btn_frame)
    
    def show_products_page(self):
        ProductsPage(self.root, self.data_manager, self.show_main_menu)
    
    def show_meal_registration_page(self):
        MealRegistrationPage(self.root, self.data_manager, self.show_main_menu)
    
    def show_employee_management(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Quản lý danh sách nhân viên")
        dialog.geometry("1200x700")
        dialog.configure(bg=COLORS['white'])
        dialog.transient(self.root)
        dialog.grab_set()
        
        screen_width = dialog.winfo_screenwidth()
        screen_height = dialog.winfo_screenheight()
        x = (screen_width - 1200) // 2
        y = (screen_height - 700) // 2
        dialog.geometry(f"1200x700+{x}+{y}")
        
        header_frame = tk.Frame(dialog, bg=COLORS['primary'], height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        
        tk.Label(
            header_frame,
            text="QUẢN LÝ DANH SÁCH NHÂN VIÊN",
            font=FONTS['header'],
            bg=COLORS['primary'],
            fg=COLORS['white']
        ).pack(pady=15)
        
        toolbar_frame = tk.Frame(dialog, bg=COLORS['light_bg'], height=60)
        toolbar_frame.pack(fill='x')
        toolbar_frame.pack_propagate(False)
        
        btn_frame = tk.Frame(toolbar_frame, bg=COLORS['light_bg'])
        btn_frame.pack(side='left', padx=10, pady=15)
        
        manage_tree = None
        manage_registrations = self.data_manager.load_meal_list()
        
        def refresh_manage_display():
            for item in manage_tree.get_children():
                manage_tree.delete(item)
            
            for idx, reg in enumerate(manage_registrations):
                tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
                
                lunch_val = float(reg['lunch'])
                dinner_val = float(reg['dinner'])
                total_val = lunch_val + dinner_val
                
                manage_tree.insert('', 'end', values=(
                    reg['stt'],
                    reg['name'],
                    reg['department'],
                    f"{lunch_val:.2f}",
                    f"{dinner_val:.2f}",
                    f"{total_val:.2f}",
                    reg['note']
                ), tags=(tag,))
        
        def add_new_item():
            add_dialog = tk.Toplevel(dialog)
            add_dialog.title("Thêm nhân viên")
            add_dialog.geometry("500x550")
            add_dialog.configure(bg=COLORS['white'])
            add_dialog.transient(dialog)
            add_dialog.grab_set()
            
            screen_w = add_dialog.winfo_screenwidth()
            screen_h = add_dialog.winfo_screenheight()
            x_pos = (screen_w - 500) // 2
            y_pos = (screen_h - 550) // 2
            add_dialog.geometry(f"500x550+{x_pos}+{y_pos}")
            
            main_frame = tk.Frame(add_dialog, bg=COLORS['white'], padx=30, pady=15)
            main_frame.pack(expand=True, fill='both')
            
            tk.Label(
                main_frame,
                text="Thêm nhân viên",
                font=FONTS['header'],
                bg=COLORS['white'],
                fg=COLORS['primary']
            ).pack(pady=(0, 12))
            
            tk.Label(main_frame, text="Họ & Tên:", font=FONTS['section'], bg=COLORS['white']).pack(anchor='w', pady=(5, 3))
            name_entry = tk.Entry(main_frame, font=('Arial', 12), width=40)
            name_entry.pack(pady=(0, 5), ipady=5)
            
            tk.Label(main_frame, text="Bộ phận:", font=FONTS['section'], bg=COLORS['white']).pack(anchor='w', pady=(5, 3))
            dept_entry = tk.Entry(main_frame, font=('Arial', 12), width=40)
            dept_entry.pack(pady=(0, 5), ipady=5)
            
            tk.Label(main_frame, text="Ăn trưa 11h30:", font=FONTS['section'], bg=COLORS['white']).pack(anchor='w', pady=(5, 5))
            lunch_var = tk.IntVar(value=1)
            lunch_frame = tk.Frame(main_frame, bg=COLORS['white'])
            lunch_frame.pack(anchor='w', pady=(0, 8))
            tk.Radiobutton(lunch_frame, text="Không ăn (0)", variable=lunch_var, value=0, font=('Arial', 12), bg=COLORS['white'], selectcolor=COLORS['light_bg']).pack(side='left', padx=(10, 30))
            tk.Radiobutton(lunch_frame, text="Có ăn (1)", variable=lunch_var, value=1, font=('Arial', 12), bg=COLORS['white'], selectcolor=COLORS['light_bg']).pack(side='left')
            
            tk.Label(main_frame, text="Ăn chiều 16h00:", font=FONTS['section'], bg=COLORS['white']).pack(anchor='w', pady=(5, 5))
            dinner_var = tk.IntVar(value=1)
            dinner_frame = tk.Frame(main_frame, bg=COLORS['white'])
            dinner_frame.pack(anchor='w', pady=(0, 8))
            tk.Radiobutton(dinner_frame, text="Không ăn (0)", variable=dinner_var, value=0, font=('Arial', 12), bg=COLORS['white'], selectcolor=COLORS['light_bg']).pack(side='left', padx=(10, 30))
            tk.Radiobutton(dinner_frame, text="Có ăn (1)", variable=dinner_var, value=1, font=('Arial', 12), bg=COLORS['white'], selectcolor=COLORS['light_bg']).pack(side='left')
            
            tk.Label(main_frame, text="Ghi chú:", font=FONTS['section'], bg=COLORS['white']).pack(anchor='w', pady=(5, 5))
            note_entry = tk.Entry(main_frame, font=('Arial', 12), width=40)
            note_entry.pack(pady=(0, 12), ipady=5)
            
            def save_new():
                name = name_entry.get().strip()
                department = dept_entry.get().strip()
                
                if not name:
                    messagebox.showwarning("Cảnh báo", "Vui lòng nhập Họ & Tên!")
                    return
                
                if manage_registrations:
                    max_stt = max(reg['stt'] for reg in manage_registrations)
                    stt = max_stt + 1
                else:
                    stt = 1
                
                lunch = float(lunch_var.get())
                dinner = float(dinner_var.get())
                
                new_reg = {
                    'stt': stt,
                    'name': name,
                    'department': department,
                    'lunch': lunch,
                    'dinner': dinner,
                    'total': lunch + dinner,
                    'note': note_entry.get().strip()
                }
                
                manage_registrations.append(new_reg)
                manage_registrations.sort(key=lambda x: x['stt'])
                refresh_manage_display()
                add_dialog.destroy()
                messagebox.showinfo("Thành công", "Đã thêm nhân viên thành công!")
            
            btn_frame_add = tk.Frame(main_frame, bg=COLORS['white'])
            btn_frame_add.pack(pady=10)
            tk.Button(btn_frame_add, text="✓ THÊM", font=('Arial', 13, 'bold'), bg=COLORS['success'], fg=COLORS['white'], width=15, height=2, command=save_new, cursor='hand2').pack(side='left', padx=10)
            tk.Button(btn_frame_add, text="✗ HỦY", font=('Arial', 13, 'bold'), bg=COLORS['secondary'], fg=COLORS['white'], width=15, height=2, command=add_dialog.destroy, cursor='hand2').pack(side='left', padx=10)
        
        def edit_item():
            item = manage_tree.selection()
            if not item:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn một dòng để sửa!")
                return
            
            values = manage_tree.item(item[0])['values']
            if not values[0]:
                return
            
            stt = values[0]
            reg = None
            for r in manage_registrations:
                if r['stt'] == stt:
                    reg = r
                    break
            
            if not reg:
                return
            
            edit_dialog = tk.Toplevel(dialog)
            edit_dialog.title("Chỉnh sửa nhân viên")
            edit_dialog.geometry("500x650")
            edit_dialog.configure(bg=COLORS['white'])
            edit_dialog.transient(dialog)
            edit_dialog.grab_set()
            
            screen_w = edit_dialog.winfo_screenwidth()
            screen_h = edit_dialog.winfo_screenheight()
            x_pos = (screen_w - 500) // 2
            y_pos = (screen_h - 650) // 2
            edit_dialog.geometry(f"500x650+{x_pos}+{y_pos}")
            
            main_frame = tk.Frame(edit_dialog, bg=COLORS['white'], padx=30, pady=20)
            main_frame.pack(expand=True, fill='both')
            
            tk.Label(main_frame, text="Chỉnh sửa nhân viên", font=FONTS['header'], bg=COLORS['white'], fg=COLORS['primary']).pack(pady=(0, 15))
            
            tk.Label(main_frame, text="STT:", font=FONTS['section'], bg=COLORS['white']).pack(anchor='w', pady=(5, 3))
            stt_entry = tk.Entry(main_frame, font=('Arial', 12), width=40)
            stt_entry.insert(0, str(reg['stt']))
            stt_entry.pack(pady=(0, 5), ipady=5)
            
            tk.Label(main_frame, text="Họ & Tên:", font=FONTS['section'], bg=COLORS['white']).pack(anchor='w', pady=(5, 3))
            name_entry = tk.Entry(main_frame, font=('Arial', 12), width=40)
            name_entry.insert(0, reg['name'])
            name_entry.pack(pady=(0, 5), ipady=5)
            
            tk.Label(main_frame, text="Bộ phận:", font=FONTS['section'], bg=COLORS['white']).pack(anchor='w', pady=(5, 3))
            dept_entry = tk.Entry(main_frame, font=('Arial', 12), width=40)
            dept_entry.insert(0, reg['department'])
            dept_entry.pack(pady=(0, 5), ipady=5)
            
            tk.Label(main_frame, text="Ăn trưa 11h30:", font=FONTS['section'], bg=COLORS['white']).pack(anchor='w', pady=(5, 5))
            lunch_var = tk.IntVar(value=int(reg['lunch']))
            lunch_frame = tk.Frame(main_frame, bg=COLORS['white'])
            lunch_frame.pack(anchor='w', pady=(0, 8))
            tk.Radiobutton(lunch_frame, text="Không ăn (0)", variable=lunch_var, value=0, font=('Arial', 12), bg=COLORS['white'], selectcolor=COLORS['light_bg']).pack(side='left', padx=(10, 30))
            tk.Radiobutton(lunch_frame, text="Có ăn (1)", variable=lunch_var, value=1, font=('Arial', 12), bg=COLORS['white'], selectcolor=COLORS['light_bg']).pack(side='left')
            
            tk.Label(main_frame, text="Ăn chiều 16h00:", font=FONTS['section'], bg=COLORS['white']).pack(anchor='w', pady=(5, 5))
            dinner_var = tk.IntVar(value=int(reg['dinner']))
            dinner_frame = tk.Frame(main_frame, bg=COLORS['white'])
            dinner_frame.pack(anchor='w', pady=(0, 8))
            tk.Radiobutton(dinner_frame, text="Không ăn (0)", variable=dinner_var, value=0, font=('Arial', 12), bg=COLORS['white'], selectcolor=COLORS['light_bg']).pack(side='left', padx=(10, 30))
            tk.Radiobutton(dinner_frame, text="Có ăn (1)", variable=dinner_var, value=1, font=('Arial', 12), bg=COLORS['white'], selectcolor=COLORS['light_bg']).pack(side='left')
            
            tk.Label(main_frame, text="Ghi chú:", font=FONTS['section'], bg=COLORS['white']).pack(anchor='w', pady=(5, 5))
            note_entry = tk.Entry(main_frame, font=('Arial', 12), width=40)
            note_entry.insert(0, reg['note'])
            note_entry.pack(pady=(0, 15), ipady=5)
            
            def save_changes():
                new_stt = stt_entry.get().strip()
                new_name = name_entry.get().strip()
                new_dept = dept_entry.get().strip()
                
                if not new_stt or not new_name:
                    messagebox.showwarning("Cảnh báo", "Vui lòng nhập STT và Họ & Tên!")
                    return
                
                try:
                    new_stt = int(new_stt)
                except ValueError:
                    messagebox.showerror("Lỗi", "STT phải là số nguyên!")
                    return
                
                if new_stt != reg['stt']:
                    for r in manage_registrations:
                        if r['stt'] == new_stt:
                            messagebox.showerror("Lỗi", f"STT {new_stt} đã tồn tại!")
                            return
                
                lunch = float(lunch_var.get())
                dinner = float(dinner_var.get())
                
                reg['stt'] = new_stt
                reg['name'] = new_name
                reg['department'] = new_dept
                reg['lunch'] = lunch
                reg['dinner'] = dinner
                reg['total'] = lunch + dinner
                reg['note'] = note_entry.get().strip()
                
                manage_registrations.sort(key=lambda x: x['stt'])
                refresh_manage_display()
                edit_dialog.destroy()
            
            btn_frame_edit = tk.Frame(main_frame, bg=COLORS['white'])
            btn_frame_edit.pack(pady=10)
            tk.Button(btn_frame_edit, text="✓ CẬP NHẬT", font=('Arial', 13, 'bold'), bg=COLORS['success'], fg=COLORS['white'], width=15, height=2, command=save_changes, cursor='hand2').pack(side='left', padx=10)
            tk.Button(btn_frame_edit, text="✗ HỦY", font=('Arial', 13, 'bold'), bg=COLORS['secondary'], fg=COLORS['white'], width=15, height=2, command=edit_dialog.destroy, cursor='hand2').pack(side='left', padx=10)
        
        def delete_item():
            item = manage_tree.selection()
            if not item:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn một dòng để xóa!")
                return
            
            values = manage_tree.item(item[0])['values']
            if not values[0]:
                return
            
            stt = values[0]
            name = values[1]
            
            result = messagebox.askyesno("Xác nhận", f"Bạn có chắc muốn xóa nhân viên:\n{name} (STT: {stt})?")
            if result:
                manage_registrations[:] = [reg for reg in manage_registrations if reg['stt'] != stt]
                refresh_manage_display()
                messagebox.showinfo("Thành công", "Đã xóa nhân viên!")
        
        def save_all_changes():
            try:
                self.data_manager.save_meal_list(manage_registrations)
                messagebox.showinfo("Thành công", "Đã lưu thay đổi vào sheet Meal_20251120!")
                dialog.destroy()
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể lưu: {str(e)}")
        
        tk.Button(btn_frame, text="➕ Thêm", font=FONTS['small_bold'], bg=COLORS['success'], fg=COLORS['white'], width=12, command=add_new_item).pack(side='left', padx=5)
        tk.Button(btn_frame, text="✏️ Sửa", font=FONTS['small_bold'], bg=COLORS['warning'], fg=COLORS['white'], width=12, command=edit_item).pack(side='left', padx=5)
        tk.Button(btn_frame, text="🗑️ Xóa", font=FONTS['small_bold'], bg=COLORS['danger'], fg=COLORS['white'], width=12, command=delete_item).pack(side='left', padx=5)
        tk.Button(btn_frame, text="💾 Lưu thay đổi", font=FONTS['small_bold'], bg=COLORS['primary'], fg=COLORS['white'], width=12, command=save_all_changes).pack(side='left', padx=5)
        tk.Button(btn_frame, text="✗ Đóng", font=FONTS['small_bold'], bg=COLORS['secondary'], fg=COLORS['white'], width=12, command=dialog.destroy).pack(side='left', padx=5)
        
        container_frame = tk.Frame(dialog, bg=COLORS['white'])
        container_frame.pack(expand=True, fill='both', padx=10, pady=10)
        
        style = ttk.Style()
        style.configure("Manage.Treeview", background=COLORS['white'], foreground="black", rowheight=30, fieldbackground=COLORS['white'], font=FONTS['table'])
        style.configure("Manage.Treeview.Heading", font=FONTS['table_header'], background=COLORS['darker'], foreground=COLORS['white'], relief='flat')
        style.map("Manage.Treeview.Heading", background=[('active', COLORS['darker'])], foreground=[('active', COLORS['white'])])
        style.map('Manage.Treeview', background=[('selected', COLORS['primary'])])
        
        scrollbar = ttk.Scrollbar(container_frame, orient='vertical')
        scrollbar.pack(side='right', fill='y')
        
        manage_tree = ttk.Treeview(container_frame, columns=('STT', 'Name', 'Department', 'Lunch', 'Dinner', 'Total', 'Note'), show='headings', yscrollcommand=scrollbar.set, style="Manage.Treeview")
        scrollbar.config(command=manage_tree.yview)
        
        manage_tree.heading('STT', text='STT', anchor='center')
        manage_tree.heading('Name', text='HỌ & TÊN', anchor='w')
        manage_tree.heading('Department', text='BỘ PHẬN', anchor='w')
        manage_tree.heading('Lunch', text='ĂN TRƯA 11h30', anchor='center')
        manage_tree.heading('Dinner', text='ĂN CHIỀU 16h00', anchor='center')
        manage_tree.heading('Total', text='PHẦN ĂN', anchor='center')
        manage_tree.heading('Note', text='GHI CHÚ', anchor='w')
        
        manage_tree.column('STT', width=60, anchor='center', stretch=False)
        manage_tree.column('Name', width=200, anchor='w', stretch=False)
        manage_tree.column('Department', width=150, anchor='w', stretch=False)
        manage_tree.column('Lunch', width=150, anchor='center', stretch=False)
        manage_tree.column('Dinner', width=150, anchor='center', stretch=False)
        manage_tree.column('Total', width=100, anchor='center', stretch=False)
        manage_tree.column('Note', width=250, anchor='w', stretch=True)
        
        manage_tree.pack(expand=True, fill='both')
        
        manage_tree.tag_configure('oddrow', background=COLORS['row_odd'])
        manage_tree.tag_configure('evenrow', background=COLORS['row_even'])
        
        manage_tree.bind('<Double-1>', lambda e: edit_item())
        
        refresh_manage_display()
    
    def show_statistics(self):
        from tkinter import ttk
        from datetime import datetime
        import calendar
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Thống kê chi phí suất ăn")
        dialog.geometry("500x400")
        dialog.configure(bg=COLORS['white'])
        dialog.transient(self.root)
        dialog.grab_set()
        
        screen_width = dialog.winfo_screenwidth()
        screen_height = dialog.winfo_screenheight()
        x = (screen_width - 500) // 2
        y = (screen_height - 400) // 2
        dialog.geometry(f"500x400+{x}+{y}")
        
        main_frame = tk.Frame(dialog, bg=COLORS['white'], padx=30, pady=20)
        main_frame.pack(expand=True, fill='both')
        
        tk.Label(
            main_frame,
            text="THỐNG KÊ CHI PHÍ SUẤT ĂN",
            font=FONTS['header'],
            bg=COLORS['white'],
            fg=COLORS['primary']
        ).pack(pady=(0, 20))
        
        tk.Label(
            main_frame,
            text="Chọn tháng/năm để thống kê:",
            font=FONTS['section'],
            bg=COLORS['white']
        ).pack(anchor='w', pady=(10, 5))
        
        date_frame = tk.Frame(main_frame, bg=COLORS['white'])
        date_frame.pack(pady=10)
        
        tk.Label(date_frame, text="Tháng:", font=FONTS['normal'], bg=COLORS['white']).pack(side='left', padx=5)
        month_var = tk.StringVar(value=str(datetime.now().month))
        month_combo = ttk.Combobox(date_frame, textvariable=month_var, values=[str(i) for i in range(1, 13)], width=5, state='readonly')
        month_combo.pack(side='left', padx=5)
        
        tk.Label(date_frame, text="Năm:", font=FONTS['normal'], bg=COLORS['white']).pack(side='left', padx=5)
        year_var = tk.StringVar(value=str(datetime.now().year))
        year_combo = ttk.Combobox(date_frame, textvariable=year_var, values=[str(i) for i in range(2020, 2031)], width=8, state='readonly')
        year_combo.pack(side='left', padx=5)
        
        progress_label = tk.Label(main_frame, text="", font=FONTS['small'], bg=COLORS['white'], fg=COLORS['success'])
        progress_label.pack(pady=10)
        
        def generate_statistics():
            try:
                month = int(month_var.get())
                year = int(year_var.get())
                
                progress_label.config(text="Đang xử lý...", fg=COLORS['warning'])
                dialog.update()
                
                data_rows = self.data_manager.get_monthly_statistics_data(month, year)
                
                dialog.destroy()
                self.show_statistics_preview(month, year, data_rows)
                
            except Exception as e:
                progress_label.config(text="", fg=COLORS['danger'])
                messagebox.showerror("Lỗi", f"Không thể tạo thống kê:\n{str(e)}")
        
        btn_frame = tk.Frame(main_frame, bg=COLORS['white'])
        btn_frame.pack(pady=20)
        
        tk.Button(
            btn_frame,
            text="✓ TẠO THỐNG KÊ",
            font=FONTS['button'],
            bg=COLORS['success'],
            fg=COLORS['white'],
            width=15,
            height=2,
            command=generate_statistics,
            cursor='hand2'
        ).pack(side='left', padx=10)
        
        tk.Button(
            btn_frame,
            text="✗ HỦY",
            font=FONTS['button'],
            bg=COLORS['secondary'],
            fg=COLORS['white'],
            width=15,
            height=2,
            command=dialog.destroy,
            cursor='hand2'
        ).pack(side='left', padx=10)
    
    def show_statistics_preview(self, month, year, data_rows):
        from tkinter import ttk
        
        preview_dialog = tk.Toplevel(self.root)
        preview_dialog.title(f"Thống kê tháng {month}/{year}")
        preview_dialog.geometry("1400x800")
        preview_dialog.configure(bg=COLORS['white'])
        preview_dialog.transient(self.root)
        preview_dialog.grab_set()
        
        screen_width = preview_dialog.winfo_screenwidth()
        screen_height = preview_dialog.winfo_screenheight()
        x = (screen_width - 1400) // 2
        y = (screen_height - 800) // 2
        preview_dialog.geometry(f"1400x800+{x}+{y}")
        
        header_frame = tk.Frame(preview_dialog, bg=COLORS['primary'], height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        
        tk.Label(
            header_frame,
            text=f"THEO DÕI CHI PHÍ SUẤT ĂN THÁNG {str(month).zfill(2)}/{year}",
            font=FONTS['header'],
            bg=COLORS['primary'],
            fg=COLORS['white']
        ).pack(pady=15)
        
        toolbar_frame = tk.Frame(preview_dialog, bg=COLORS['light_bg'], height=70)
        toolbar_frame.pack(fill='x')
        toolbar_frame.pack_propagate(False)
        
        btn_frame = tk.Frame(toolbar_frame, bg=COLORS['light_bg'])
        btn_frame.pack(pady=15)
        
        def save_to_excel():
            try:
                result = self.data_manager.generate_monthly_statistics(month, year, data_rows)
                messagebox.showinfo("Thành công", result)
                preview_dialog.destroy()
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể lưu file:\n{str(e)}")
        
        tk.Button(
            btn_frame,
            text="💾 LƯU VÀO EXCEL",
            font=FONTS['button'],
            bg=COLORS['success'],
            fg=COLORS['white'],
            width=20,
            height=2,
            command=save_to_excel,
            cursor='hand2'
        ).pack(side='left', padx=10)
        
        tk.Button(
            btn_frame,
            text="✗ HỦY",
            font=FONTS['button'],
            bg=COLORS['secondary'],
            fg=COLORS['white'],
            width=15,
            height=2,
            command=preview_dialog.destroy,
            cursor='hand2'
        ).pack(side='left', padx=10)
        
        container_frame = tk.Frame(preview_dialog, bg=COLORS['white'])
        container_frame.pack(expand=True, fill='both', padx=10, pady=10)
        
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Stats.Treeview", background=COLORS['white'], foreground="black", rowheight=30, fieldbackground=COLORS['white'], font=FONTS['table'])
        style.configure("Stats.Treeview.Heading", font=('Arial', 13, 'bold'), background=COLORS['darker'], foreground="white", relief='raised', borderwidth=2, padding=[0, 30])
        style.map('Stats.Treeview.Heading', background=[('active', COLORS['darker'])], foreground=[('active', 'white')])
        style.map('Stats.Treeview', background=[('selected', COLORS['primary'])])
        
        scrollbar_y = ttk.Scrollbar(container_frame, orient='vertical')
        scrollbar_y.pack(side='right', fill='y')
        
        scrollbar_x = ttk.Scrollbar(container_frame, orient='horizontal')
        scrollbar_x.pack(side='bottom', fill='x')
        
        stats_tree = ttk.Treeview(
            container_frame,
            columns=('Day', 'DayName', 'TotalMeals', 'TotalCost', 'AvgPrice', 'WeekAvg', 'Note'),
            show='headings',
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set,
            style="Stats.Treeview"
        )
        
        scrollbar_y.config(command=stats_tree.yview)
        scrollbar_x.config(command=stats_tree.xview)
        
        stats_tree.heading('Day', text='Ngày', anchor='center')
        stats_tree.heading('DayName', text='Thứ', anchor='center')
        stats_tree.heading('TotalMeals', text='Số phần ăn\nTrưa + Chiều', anchor='center')
        stats_tree.heading('TotalCost', text='Chi phí\nthực phẩm', anchor='center')
        stats_tree.heading('AvgPrice', text='Đơn giá\nmỗi suất ăn', anchor='center')
        stats_tree.heading('WeekAvg', text='Trung bình suất\năn của tuần', anchor='center')
        stats_tree.heading('Note', text='Ghi chú\n(nhấp đôi để nhập)', anchor='center')
        
        stats_tree.column('Day', width=70, anchor='center', stretch=False)
        stats_tree.column('DayName', width=80, anchor='center', stretch=False)
        stats_tree.column('TotalMeals', width=150, anchor='center', stretch=False)
        stats_tree.column('TotalCost', width=180, anchor='center', stretch=False)
        stats_tree.column('AvgPrice', width=180, anchor='center', stretch=False)
        stats_tree.column('WeekAvg', width=180, anchor='center', stretch=False)
        stats_tree.column('Note', width=200, anchor='center', stretch=True)
        
        stats_tree.pack(expand=True, fill='both')
        
        for idx, row_data in enumerate(data_rows):
            day = row_data['day']
            day_name = row_data['day_name']
            total_meals = row_data['total_meals']
            total_cost = row_data['total_cost']
            avg_price = row_data['avg_price']
            week_avg = row_data.get('week_avg', 0)
            note = row_data['note']
            weekday = row_data['weekday']
            
            tag = 'weekend' if weekday in [5, 6] else ('evenrow' if idx % 2 == 0 else 'oddrow')
            
            if total_meals == 0:
                stats_tree.insert('', 'end', values=(
                    day,
                    day_name,
                    0,
                    '-',
                    '',
                    '',
                    note
                ), tags=(tag, 'holiday'))
            else:
                cost_str = f'{int(total_cost):,}' if total_cost > 0 else '-'
                avg_str = f'{int(avg_price):,}' if avg_price > 0 else ''
                week_avg_str = f'{int(week_avg):,}' if week_avg > 0 else ''
                
                stats_tree.insert('', 'end', values=(
                    day,
                    day_name,
                    int(total_meals),
                    cost_str,
                    avg_str,
                    week_avg_str,
                    note
                ), tags=(tag,))
        
        stats_tree.tag_configure('oddrow', background=COLORS['row_odd'])
        stats_tree.tag_configure('evenrow', background=COLORS['row_even'])
        stats_tree.tag_configure('weekend', background='#FFFF99')
        stats_tree.tag_configure('holiday', foreground='#FF0000', font=('Arial', 11, 'italic'))
        
        def edit_note(event):
            selection = stats_tree.selection()
            if not selection:
                return
            
            item = selection[0]
            item_index = stats_tree.index(item)
            current_note = data_rows[item_index]['note']
            
            note_dialog = tk.Toplevel(preview_dialog)
            note_dialog.title("Chỉnh sửa ghi chú")
            note_dialog.geometry("400x200")
            note_dialog.configure(bg=COLORS['white'])
            note_dialog.transient(preview_dialog)
            note_dialog.grab_set()
            
            screen_width = note_dialog.winfo_screenwidth()
            screen_height = note_dialog.winfo_screenheight()
            x = (screen_width - 400) // 2
            y = (screen_height - 200) // 2
            note_dialog.geometry(f"400x200+{x}+{y}")
            
            tk.Label(
                note_dialog,
                text=f"Ghi chú cho ngày {data_rows[item_index]['day']}/{month}/{year}:",
                font=FONTS['section'],
                bg=COLORS['white']
            ).pack(pady=10)
            
            note_entry = tk.Entry(note_dialog, font=FONTS['normal'], width=40)
            note_entry.insert(0, current_note)
            note_entry.pack(pady=10)
            note_entry.focus()
            
            def save_note():
                new_note = note_entry.get().strip()
                data_rows[item_index]['note'] = new_note
                
                values = list(stats_tree.item(item, 'values'))
                values[6] = new_note
                stats_tree.item(item, values=values)
                
                note_dialog.destroy()
            
            btn_frame = tk.Frame(note_dialog, bg=COLORS['white'])
            btn_frame.pack(pady=10)
            
            tk.Button(
                btn_frame,
                text="✓ LƯU",
                font=FONTS['button'],
                bg=COLORS['success'],
                fg=COLORS['white'],
                width=10,
                command=save_note,
                cursor='hand2'
            ).pack(side='left', padx=5)
            
            tk.Button(
                btn_frame,
                text="✗ HỦY",
                font=FONTS['button'],
                bg=COLORS['secondary'],
                fg=COLORS['white'],
                width=10,
                command=note_dialog.destroy,
                cursor='hand2'
            ).pack(side='left', padx=5)
            
            note_entry.bind('<Return>', lambda e: save_note())
        
        stats_tree.bind('<Double-Button-1>', edit_note)
    
    def show_menu_management(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Quản lý Menu")
        dialog.geometry("900x600")
        dialog.configure(bg=COLORS['white'])
        dialog.transient(self.root)
        dialog.grab_set()
        
        screen_width = dialog.winfo_screenwidth()
        screen_height = dialog.winfo_screenheight()
        x = (screen_width - 900) // 2
        y = (screen_height - 600) // 2
        dialog.geometry(f"900x600+{x}+{y}")
        
        header_frame = tk.Frame(dialog, bg=COLORS['primary'], height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        
        tk.Label(
            header_frame,
            text="QUẢN LÝ MENU",
            font=FONTS['header'],
            bg=COLORS['primary'],
            fg=COLORS['white']
        ).pack(pady=15)
        
        toolbar_frame = tk.Frame(dialog, bg=COLORS['light_bg'], height=60)
        toolbar_frame.pack(fill='x')
        toolbar_frame.pack_propagate(False)
        
        btn_frame = tk.Frame(toolbar_frame, bg=COLORS['light_bg'])
        btn_frame.pack(side='left', padx=10, pady=15)
        
        menu_tree = None
        
        def refresh_menu_list():
            for item in menu_tree.get_children():
                menu_tree.delete(item)
            
            menu_names = self.data_manager.get_all_menu_names()
            for idx, menu_name in enumerate(menu_names):
                tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
                items = self.data_manager.get_menu_items(menu_name)
                total_items = len(items.get('lunch', [])) + len(items.get('dinner', []))
                menu_tree.insert('', 'end', values=(idx + 1, menu_name, total_items), tags=(tag,))
        
        def create_new_menu():
            self.show_menu_editor(dialog, None, refresh_menu_list)
        
        def edit_menu():
            selected = menu_tree.selection()
            if not selected:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn một menu để sửa!")
                return
            
            values = menu_tree.item(selected[0])['values']
            menu_name = values[1]
            self.show_menu_editor(dialog, menu_name, refresh_menu_list)
        
        def delete_menu():
            selected = menu_tree.selection()
            if not selected:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn một menu để xóa!")
                return
            
            values = menu_tree.item(selected[0])['values']
            menu_name = values[1]
            
            if messagebox.askyesno("Xác nhận", f"Bạn có chắc muốn xóa menu '{menu_name}'?"):
                self.data_manager.delete_menu(menu_name)
                refresh_menu_list()
                messagebox.showinfo("Thành công", f"Đã xóa menu '{menu_name}'!")
        
        def view_menu():
            selected = menu_tree.selection()
            if not selected:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn một menu để xem!")
                return
            
            values = menu_tree.item(selected[0])['values']
            menu_name = values[1]
            self.show_menu_viewer(dialog, menu_name)
        
        tk.Button(
            btn_frame,
            text="➕ Tạo menu mới",
            font=FONTS['small_bold'],
            bg=COLORS['success'],
            fg=COLORS['white'],
            width=15,
            command=create_new_menu
        ).pack(side='left', padx=5)
        
        tk.Button(
            btn_frame,
            text="✏️ Sửa menu",
            font=FONTS['small_bold'],
            bg=COLORS['info'],
            fg=COLORS['white'],
            width=15,
            command=edit_menu
        ).pack(side='left', padx=5)
        
        tk.Button(
            btn_frame,
            text="🗑️ Xóa menu",
            font=FONTS['small_bold'],
            bg=COLORS['danger'],
            fg=COLORS['white'],
            width=15,
            command=delete_menu
        ).pack(side='left', padx=5)
        
        tk.Button(
            btn_frame,
            text="👁️ Xem menu",
            font=FONTS['small_bold'],
            bg=COLORS['warning'],
            fg=COLORS['white'],
            width=15,
            command=view_menu
        ).pack(side='left', padx=5)
        
        container_frame = tk.Frame(dialog, bg=COLORS['white'])
        container_frame.pack(expand=True, fill='both', padx=20, pady=10)
        
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Menu.Treeview", background=COLORS['white'], foreground="black", rowheight=30, fieldbackground=COLORS['white'], font=FONTS['table'])
        style.configure("Menu.Treeview.Heading", font=FONTS['table_header'], background=COLORS['darker'], foreground=COLORS['white'], relief='flat')
        style.map("Menu.Treeview.Heading", background=[('active', COLORS['darker'])], foreground=[('active', COLORS['white'])])
        style.map('Menu.Treeview', background=[('selected', COLORS['primary']), ('!selected', '')], foreground=[('selected', 'white')])
        
        scrollbar = ttk.Scrollbar(container_frame, orient='vertical')
        scrollbar.pack(side='right', fill='y')
        
        menu_tree = ttk.Treeview(
            container_frame,
            columns=('STT', 'MenuName', 'Items'),
            show='headings',
            yscrollcommand=scrollbar.set,
            style="Menu.Treeview"
        )
        scrollbar.config(command=menu_tree.yview)
        
        menu_tree.heading('STT', text='STT', anchor='center')
        menu_tree.heading('MenuName', text='Tên Menu', anchor='w')
        menu_tree.heading('Items', text='Số món', anchor='center')
        
        menu_tree.column('STT', width=60, anchor='center', stretch=False)
        menu_tree.column('MenuName', width=400, anchor='w', stretch=True)
        menu_tree.column('Items', width=100, anchor='center', stretch=False)
        
        menu_tree.pack(expand=True, fill='both')
        
        menu_tree.tag_configure('oddrow', background=COLORS['row_odd'], foreground='black')
        menu_tree.tag_configure('evenrow', background=COLORS['row_even'], foreground='black')
        
        refresh_menu_list()
        
        btn_close_frame = tk.Frame(dialog, bg=COLORS['white'])
        btn_close_frame.pack(pady=10)
        tk.Button(
            btn_close_frame,
            text="✗ Đóng",
            font=FONTS['small_bold'],
            bg=COLORS['secondary'],
            fg=COLORS['white'],
            width=12,
            command=dialog.destroy
        ).pack(padx=5)
    
    def show_menu_editor(self, parent, menu_name, on_save_callback):
        editor_dialog = tk.Toplevel(parent)
        title = "Sửa Menu" if menu_name else "Tạo Menu Mới"
        editor_dialog.title(title)
        editor_dialog.geometry("1300x700")
        editor_dialog.configure(bg=COLORS['white'])
        editor_dialog.transient(parent)
        editor_dialog.grab_set()
        
        screen_width = editor_dialog.winfo_screenwidth()
        screen_height = editor_dialog.winfo_screenheight()
        x = (screen_width - 1300) // 2
        y = (screen_height - 700) // 2
        editor_dialog.geometry(f"1300x700+{x}+{y}")
        
        header_frame = tk.Frame(editor_dialog, bg=COLORS['primary'], height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        
        tk.Label(
            header_frame,
            text=title.upper(),
            font=FONTS['header'],
            bg=COLORS['primary'],
            fg=COLORS['white']
        ).pack(pady=15)
        
        menu_name_frame = tk.Frame(editor_dialog, bg=COLORS['light_bg'])
        menu_name_frame.pack(fill='x', padx=20, pady=10)
        
        tk.Label(
            menu_name_frame,
            text="Tháng:",
            font=FONTS['normal'],
            bg=COLORS['light_bg']
        ).pack(side='left', padx=10)
        
        month_combo = ttk.Combobox(menu_name_frame, values=list(range(1, 13)), width=8, font=FONTS['normal'], state='readonly')
        month_combo.pack(side='left', padx=5)
        
        tk.Label(
            menu_name_frame,
            text="Tuần:",
            font=FONTS['normal'],
            bg=COLORS['light_bg']
        ).pack(side='left', padx=10)
        
        week_combo = ttk.Combobox(menu_name_frame, values=list(range(1, 3)), width=8, font=FONTS['normal'], state='readonly')
        week_combo.pack(side='left', padx=5)
        
        tk.Label(
            menu_name_frame,
            text="Thứ:",
            font=FONTS['normal'],
            bg=COLORS['light_bg']
        ).pack(side='left', padx=10)
        
        day_combo = ttk.Combobox(menu_name_frame, values=list(range(2, 7)), width=8, font=FONTS['normal'], state='readonly')
        day_combo.pack(side='left', padx=5)
        
        if menu_name:
            parts = menu_name.split('-')
            if len(parts) == 3 and parts[0].startswith('T'):
                month_combo.set(parts[0][1:])
                week_combo.set(parts[1])
                day_combo.set(parts[2])
                month_combo.config(state='disabled')
                week_combo.config(state='disabled')
                day_combo.config(state='disabled')
        
        products_data = self.data_manager.load_data("Products")
        fruits_data = self.data_manager.load_data("Fruits")
        all_products = {}
        for row in products_data:
            if row[1]:
                all_products[row[1]] = {'unit': row[2], 'price': row[3]}
        for row in fruits_data:
            if row[1]:
                all_products[row[1]] = {'unit': row[2], 'price': row[3]}
        product_names = list(all_products.keys())
        
        lunch_items = []
        dinner_items = []
        if menu_name:
            existing_items = self.data_manager.get_menu_items(menu_name)
            lunch_items = existing_items.get('lunch', [])
            dinner_items = existing_items.get('dinner', [])
        
        main_container = tk.Frame(editor_dialog, bg=COLORS['white'])
        main_container.pack(expand=True, fill='both', padx=10, pady=10)
        
        lunch_frame = tk.Frame(main_container, bg=COLORS['light_bg'], relief='ridge', bd=2)
        lunch_frame.pack(side='left', expand=True, fill='both', padx=5)
        
        tk.Label(lunch_frame, text="THỰC ĐƠN TRƯA", font=FONTS['title'], bg=COLORS['info'], fg=COLORS['white']).pack(fill='x', pady=5)
        
        dinner_frame = tk.Frame(main_container, bg=COLORS['light_bg'], relief='ridge', bd=2)
        dinner_frame.pack(side='right', expand=True, fill='both', padx=5)
        
        tk.Label(dinner_frame, text="THỰC ĐƠN CHIỀU", font=FONTS['title'], bg=COLORS['warning'], fg=COLORS['white']).pack(fill='x', pady=5)
        
        def create_meal_section(container, items_list, section_name):
            edit_index = None
            
            form_frame = tk.Frame(container, bg=COLORS['light_bg'])
            form_frame.pack(fill='x', padx=10, pady=10)
            
            tk.Label(form_frame, text="Tên thực phẩm:", font=FONTS['small'], bg=COLORS['light_bg']).grid(row=0, column=0, sticky='e', padx=5, pady=3)
            product_combo = ttk.Combobox(form_frame, values=product_names, width=20, font=FONTS['small'])
            product_combo.grid(row=0, column=1, padx=5, pady=3)
            
            tk.Label(form_frame, text="Đơn vị:", font=FONTS['small'], bg=COLORS['light_bg']).grid(row=1, column=0, sticky='e', padx=5, pady=3)
            unit_label = tk.Label(form_frame, text="", font=FONTS['small'], bg=COLORS['light_bg'], width=10, anchor='w')
            unit_label.grid(row=1, column=1, padx=5, pady=3, sticky='w')
            
            tk.Label(form_frame, text="Định mức:", font=FONTS['small'], bg=COLORS['light_bg']).grid(row=2, column=0, sticky='e', padx=5, pady=3)
            qty_entry = tk.Entry(form_frame, width=12, font=FONTS['small'])
            qty_entry.grid(row=2, column=1, padx=5, pady=3, sticky='w')
            
            tk.Label(form_frame, text="Đơn giá:", font=FONTS['small'], bg=COLORS['light_bg']).grid(row=3, column=0, sticky='e', padx=5, pady=3)
            price_entry = tk.Entry(form_frame, width=12, font=FONTS['small'])
            price_entry.insert(0, "0")
            price_entry.grid(row=3, column=1, padx=5, pady=3, sticky='w')
            
            def on_product_change(event=None):
                product_name = product_combo.get()
                if product_name in all_products:
                    product_info = all_products[product_name]
                    unit = product_info['unit']
                    price = product_info['price']
                    
                    unit_label.config(text=unit)
                    qty_entry.delete(0, tk.END)
                    qty_entry.insert(0, "1")
                    price_entry.delete(0, tk.END)
                    price_entry.insert(0, f"{price:,.0f}")
            
            def on_keyrelease(event):
                value = event.widget.get()
                if value == '':
                    product_combo['values'] = product_names
                else:
                    filtered = [item for item in product_names if value.lower() in item.lower()]
                    product_combo['values'] = filtered
            
            product_combo.bind('<<ComboboxSelected>>', on_product_change)
            product_combo.bind('<KeyRelease>', on_keyrelease)
            
            def refresh_display():
                for item in items_tree.get_children():
                    items_tree.delete(item)
                
                for idx, item in enumerate(items_list):
                    tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
                    items_tree.insert('', 'end', values=(
                        idx + 1,
                        item['product_name'],
                        f"{item['qty']:.2f}",
                        item['unit'],
                        f"{item['price']:,.0f}"
                    ), tags=(tag,))
            
            def clear_form():
                nonlocal edit_index
                edit_index = None
                product_combo.set('')
                qty_entry.delete(0, tk.END)
                unit_label.config(text="")
                price_entry.delete(0, tk.END)
                price_entry.insert(0, "0")
                btn_add.config(text="➕ Thêm")
            
            def load_item(event):
                nonlocal edit_index
                selected = items_tree.selection()
                if not selected:
                    return
                
                values = items_tree.item(selected[0])['values']
                edit_index = values[0] - 1
                
                if edit_index < len(items_list):
                    item = items_list[edit_index]
                    product_combo.set(item['product_name'])
                    unit_label.config(text=item['unit'])
                    qty_entry.delete(0, tk.END)
                    qty_entry.insert(0, str(item['qty']))
                    price_entry.delete(0, tk.END)
                    price_entry.insert(0, f"{item['price']:,.0f}")
                    btn_add.config(text="✏️ Sửa")
            
            def add_or_update():
                nonlocal edit_index
                product_name = product_combo.get()
                if not product_name:
                    messagebox.showwarning("Cảnh báo", "Vui lòng chọn tên thực phẩm!")
                    return
                
                try:
                    qty = float(qty_entry.get())
                    unit = unit_label.cget("text")
                    price = float(price_entry.get().replace(",", ""))
                    
                    item_data = {
                        'product_name': product_name,
                        'qty': qty,
                        'unit': unit,
                        'price': price
                    }
                    
                    if edit_index is not None:
                        items_list[edit_index] = item_data
                    else:
                        items_list.append(item_data)
                    
                    refresh_display()
                    clear_form()
                    
                except ValueError:
                    messagebox.showerror("Lỗi", "Vui lòng nhập đúng định dạng số!")
            
            def delete_item():
                nonlocal edit_index
                selected = items_tree.selection()
                if not selected:
                    messagebox.showwarning("Cảnh báo", "Vui lòng chọn một dòng để xóa!")
                    return
                
                values = items_tree.item(selected[0])['values']
                idx = values[0] - 1
                
                if idx < len(items_list):
                    items_list.pop(idx)
                    if edit_index is not None and edit_index >= idx:
                        edit_index = None
                    refresh_display()
                    clear_form()
            
            btn_frame = tk.Frame(form_frame, bg=COLORS['light_bg'])
            btn_frame.grid(row=4, column=0, columnspan=2, pady=10)
            
            btn_add = tk.Button(btn_frame, text="➕ Thêm", font=FONTS['small_bold'], bg=COLORS['success'], fg=COLORS['white'], width=10, command=add_or_update)
            btn_add.pack(side='left', padx=3)
            
            tk.Button(btn_frame, text="🗑️ Xóa", font=FONTS['small_bold'], bg=COLORS['danger'], fg=COLORS['white'], width=10, command=delete_item).pack(side='left', padx=3)
            
            tk.Button(btn_frame, text="✗ Hủy", font=FONTS['small_bold'], bg=COLORS['secondary'], fg=COLORS['white'], width=10, command=clear_form).pack(side='left', padx=3)
            
            table_frame = tk.Frame(container, bg=COLORS['white'])
            table_frame.pack(expand=True, fill='both', padx=10, pady=5)
            
            scrollbar = ttk.Scrollbar(table_frame, orient='vertical')
            scrollbar.pack(side='right', fill='y')
            
            items_tree = ttk.Treeview(
                table_frame,
                columns=('STT', 'Product', 'Qty', 'Unit', 'Price'),
                show='headings',
                yscrollcommand=scrollbar.set,
                style=f"{section_name}.Treeview"
            )
            scrollbar.config(command=items_tree.yview)
            
            items_tree.heading('STT', text='STT', anchor='center')
            items_tree.heading('Product', text='Tên thực phẩm', anchor='w')
            items_tree.heading('Qty', text='Định mức', anchor='center')
            items_tree.heading('Unit', text='ĐV', anchor='center')
            items_tree.heading('Price', text='Đơn giá', anchor='center')
            
            items_tree.column('STT', width=40, anchor='center', stretch=False)
            items_tree.column('Product', width=180, anchor='w', stretch=True)
            items_tree.column('Qty', width=70, anchor='center', stretch=False)
            items_tree.column('Unit', width=50, anchor='center', stretch=False)
            items_tree.column('Price', width=90, anchor='center', stretch=False)
            
            items_tree.pack(expand=True, fill='both')
            
            items_tree.tag_configure('oddrow', background=COLORS['row_odd'], foreground='black')
            items_tree.tag_configure('evenrow', background=COLORS['row_even'], foreground='black')
            
            items_tree.bind('<ButtonRelease-1>', load_item)
            
            refresh_display()
        
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Lunch.Treeview", background=COLORS['white'], foreground="black", rowheight=25, fieldbackground=COLORS['white'], font=FONTS['small'])
        style.configure("Lunch.Treeview.Heading", font=FONTS['small_bold'], background=COLORS['info'], foreground=COLORS['white'], relief='flat')
        style.map("Lunch.Treeview.Heading", background=[('active', COLORS['info'])], foreground=[('active', COLORS['white'])])
        style.map('Lunch.Treeview', background=[('selected', COLORS['primary'])], foreground=[('selected', 'white')])
        
        style.configure("Dinner.Treeview", background=COLORS['white'], foreground="black", rowheight=25, fieldbackground=COLORS['white'], font=FONTS['small'])
        style.configure("Dinner.Treeview.Heading", font=FONTS['small_bold'], background=COLORS['warning'], foreground=COLORS['white'], relief='flat')
        style.map("Dinner.Treeview.Heading", background=[('active', COLORS['warning'])], foreground=[('active', COLORS['white'])])
        style.map('Dinner.Treeview', background=[('selected', COLORS['primary'])], foreground=[('selected', 'white')])
        
        create_meal_section(lunch_frame, lunch_items, "Lunch")
        create_meal_section(dinner_frame, dinner_items, "Dinner")
        
        def save_menu():
            month = month_combo.get()
            week = week_combo.get()
            day = day_combo.get()
            
            if not month or not week or not day:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn đầy đủ Tháng, Tuần và Thứ!")
                return
            
            new_menu_name = f"T{month}-{week}-{day}"
            
            if not lunch_items and not dinner_items:
                messagebox.showwarning("Cảnh báo", "Vui lòng thêm ít nhất một món vào menu!")
                return
            
            if new_menu_name.lower() != (menu_name.lower() if menu_name else ""):
                existing_menu = self.data_manager.get_menu_items(new_menu_name)
                if existing_menu and (existing_menu.get('lunch') or existing_menu.get('dinner')):
                    response = messagebox.askyesno(
                        "Menu đã tồn tại",
                        f"Menu '{new_menu_name}' đã tồn tại.\nBạn có muốn thay thế menu cũ không?"
                    )
                    if not response:
                        return
            
            self.data_manager.save_menu(new_menu_name, lunch_items, dinner_items)
            messagebox.showinfo("Thành công", f"Đã lưu menu '{new_menu_name}'!")
            on_save_callback()
            editor_dialog.destroy()
        
        btn_bottom_frame = tk.Frame(editor_dialog, bg=COLORS['white'])
        btn_bottom_frame.pack(pady=10)
        
        save_btn_text = "💾 Cập nhật" if menu_name else "💾 Lưu Menu"
        tk.Button(
            btn_bottom_frame,
            text=save_btn_text,
            font=FONTS['button'],
            bg=COLORS['success'],
            fg=COLORS['white'],
            width=15,
            height=2,
            command=save_menu
        ).pack(side='left', padx=10)
        
        tk.Button(
            btn_bottom_frame,
            text="✗ Hủy",
            font=FONTS['button'],
            bg=COLORS['secondary'],
            fg=COLORS['white'],
            width=15,
            height=2,
            command=editor_dialog.destroy
        ).pack(side='left', padx=10)
    
    def show_menu_viewer(self, parent, menu_name):
        viewer_dialog = tk.Toplevel(parent)
        viewer_dialog.title(f"Chi tiết Menu: {menu_name}")
        viewer_dialog.geometry("1300x600")
        viewer_dialog.configure(bg=COLORS['white'])
        viewer_dialog.transient(parent)
        viewer_dialog.grab_set()
        
        screen_width = viewer_dialog.winfo_screenwidth()
        screen_height = viewer_dialog.winfo_screenheight()
        x = (screen_width - 1300) // 2
        y = (screen_height - 600) // 2
        viewer_dialog.geometry(f"1300x600+{x}+{y}")
        
        header_frame = tk.Frame(viewer_dialog, bg=COLORS['primary'], height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        
        tk.Label(
            header_frame,
            text=f"CHI TIẾT MENU: {menu_name.upper()}",
            font=FONTS['header'],
            bg=COLORS['primary'],
            fg=COLORS['white']
        ).pack(pady=15)
        
        menu_items = self.data_manager.get_menu_items(menu_name)
        lunch_items = menu_items.get('lunch', [])
        dinner_items = menu_items.get('dinner', [])
        
        main_container = tk.Frame(viewer_dialog, bg=COLORS['white'])
        main_container.pack(expand=True, fill='both', padx=10, pady=10)
        
        lunch_frame = tk.Frame(main_container, bg=COLORS['white'], relief='ridge', bd=2)
        lunch_frame.pack(side='left', expand=True, fill='both', padx=5)
        
        tk.Label(lunch_frame, text="THỰC ĐƠN TRƯA", font=FONTS['title'], bg=COLORS['info'], fg=COLORS['white']).pack(fill='x', pady=5)
        
        dinner_frame = tk.Frame(main_container, bg=COLORS['white'], relief='ridge', bd=2)
        dinner_frame.pack(side='right', expand=True, fill='both', padx=5)
        
        tk.Label(dinner_frame, text="THỰC ĐƠN CHIỀU", font=FONTS['title'], bg=COLORS['warning'], fg=COLORS['white']).pack(fill='x', pady=5)
        
        def create_viewer_section(container, items_list, section_name):
            table_frame = tk.Frame(container, bg=COLORS['white'])
            table_frame.pack(expand=True, fill='both', padx=10, pady=10)
            
            scrollbar = ttk.Scrollbar(table_frame, orient='vertical')
            scrollbar.pack(side='right', fill='y')
            
            items_tree = ttk.Treeview(
                table_frame,
                columns=('STT', 'Product', 'Qty', 'Unit', 'Price', 'Total'),
                show='headings',
                yscrollcommand=scrollbar.set,
                style=f"Viewer{section_name}.Treeview"
            )
            scrollbar.config(command=items_tree.yview)
            
            items_tree.heading('STT', text='STT', anchor='center')
            items_tree.heading('Product', text='Tên thực phẩm', anchor='w')
            items_tree.heading('Qty', text='Định mức', anchor='center')
            items_tree.heading('Unit', text='ĐV', anchor='center')
            items_tree.heading('Price', text='Đơn giá', anchor='center')
            items_tree.heading('Total', text='Thành tiền', anchor='center')
            
            items_tree.column('STT', width=40, anchor='center', stretch=False)
            items_tree.column('Product', width=180, anchor='w', stretch=True)
            items_tree.column('Qty', width=70, anchor='center', stretch=False)
            items_tree.column('Unit', width=50, anchor='center', stretch=False)
            items_tree.column('Price', width=80, anchor='center', stretch=False)
            items_tree.column('Total', width=100, anchor='center', stretch=False)
            
            items_tree.pack(expand=True, fill='both')
            
            items_tree.tag_configure('oddrow', background=COLORS['row_odd'], foreground='black')
            items_tree.tag_configure('evenrow', background=COLORS['row_even'], foreground='black')
            items_tree.tag_configure('total_row', background='#ffffcc', foreground='black', font=FONTS['table_header'])
            
            total_price = 0
            for idx, item in enumerate(items_list):
                tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
                item_total = item['qty'] * item['price']
                total_price += item_total
                items_tree.insert('', 'end', values=(
                    idx + 1,
                    item['product_name'],
                    f"{item['qty']:.2f}",
                    item['unit'],
                    f"{item['price']:,.0f}",
                    f"{item_total:,.0f}"
                ), tags=(tag,))
            
            items_tree.insert('', 'end', values=(
                '',
                'TỔNG CỘNG (1 phần)',
                '',
                '',
                '',
                f"{total_price:,.0f}"
            ), tags=('total_row',))
        
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("ViewerLunch.Treeview", background=COLORS['white'], foreground="black", rowheight=25, fieldbackground=COLORS['white'], font=FONTS['small'])
        style.configure("ViewerLunch.Treeview.Heading", font=FONTS['small_bold'], background=COLORS['info'], foreground=COLORS['white'], relief='flat')
        style.map("ViewerLunch.Treeview.Heading", background=[('active', COLORS['info'])], foreground=[('active', COLORS['white'])])
        style.map('ViewerLunch.Treeview', background=[('selected', COLORS['primary'])], foreground=[('selected', 'white')])
        
        style.configure("ViewerDinner.Treeview", background=COLORS['white'], foreground="black", rowheight=25, fieldbackground=COLORS['white'], font=FONTS['small'])
        style.configure("ViewerDinner.Treeview.Heading", font=FONTS['small_bold'], background=COLORS['warning'], foreground=COLORS['white'], relief='flat')
        style.map("ViewerDinner.Treeview.Heading", background=[('active', COLORS['warning'])], foreground=[('active', COLORS['white'])])
        style.map('ViewerDinner.Treeview', background=[('selected', COLORS['primary'])], foreground=[('selected', 'white')])
        
        create_viewer_section(lunch_frame, lunch_items, "Lunch")
        create_viewer_section(dinner_frame, dinner_items, "Dinner")
        
        btn_frame = tk.Frame(viewer_dialog, bg=COLORS['white'])
        btn_frame.pack(pady=10)
        
        tk.Button(
            btn_frame,
            text="✗ Đóng",
            font=FONTS['small_bold'],
            bg=COLORS['secondary'],
            fg=COLORS['white'],
            width=12,
            command=viewer_dialog.destroy
        ).pack(padx=5)
    
    def show_order_management(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Đặt Hàng Thực Phẩm")
        dialog.geometry("1400x750")
        dialog.configure(bg=COLORS['white'])
        dialog.transient(self.root)
        dialog.grab_set()
        
        screen_width = dialog.winfo_screenwidth()
        screen_height = dialog.winfo_screenheight()
        x = (screen_width - 1400) // 2
        y = (screen_height - 750) // 2
        dialog.geometry(f"1400x750+{x}+{y}")
        
        header_frame = tk.Frame(dialog, bg=COLORS['primary'], height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        
        tk.Label(
            header_frame,
            text="ĐẶT HÀNG THỰC PHẨM",
            font=FONTS['header'],
            bg=COLORS['primary'],
            fg=COLORS['white']
        ).pack(pady=15)
        
        input_frame = tk.Frame(dialog, bg=COLORS['light_bg'])
        input_frame.pack(fill='x', padx=20, pady=15)
        
        tk.Label(input_frame, text="Tháng:", font=FONTS['normal'], bg=COLORS['light_bg']).grid(row=0, column=0, sticky='e', padx=10, pady=5)
        month_combo = ttk.Combobox(input_frame, values=list(range(1, 13)), width=10, font=FONTS['normal'], state='readonly')
        month_combo.grid(row=0, column=1, padx=10, pady=5, sticky='w')
        
        tk.Label(input_frame, text="Tuần:", font=FONTS['normal'], bg=COLORS['light_bg']).grid(row=0, column=2, sticky='e', padx=10, pady=5)
        week_combo = ttk.Combobox(input_frame, values=list(range(1, 5)), width=10, font=FONTS['normal'], state='readonly')
        week_combo.grid(row=0, column=3, padx=10, pady=5, sticky='w')
        
        tk.Label(input_frame, text="Thứ:", font=FONTS['normal'], bg=COLORS['light_bg']).grid(row=0, column=4, sticky='e', padx=10, pady=5)
        day_combo = ttk.Combobox(input_frame, values=list(range(2, 7)), width=10, font=FONTS['normal'], state='readonly')
        day_combo.grid(row=0, column=5, padx=10, pady=5, sticky='w')
        
        tk.Label(input_frame, text="Số người ăn trưa:", font=FONTS['normal'], bg=COLORS['light_bg']).grid(row=1, column=0, sticky='e', padx=10, pady=5)
        lunch_people_entry = tk.Entry(input_frame, width=12, font=FONTS['normal'])
        lunch_people_entry.grid(row=1, column=1, padx=10, pady=5, sticky='w')
        
        tk.Label(input_frame, text="Số người ăn chiều:", font=FONTS['normal'], bg=COLORS['light_bg']).grid(row=1, column=2, sticky='e', padx=10, pady=5)
        dinner_people_entry = tk.Entry(input_frame, width=12, font=FONTS['normal'])
        dinner_people_entry.grid(row=1, column=3, padx=10, pady=5, sticky='w')
        
        today = datetime.now()
        
        tk.Label(input_frame, text="Ngày đặt hàng:", font=FONTS['normal'], bg=COLORS['light_bg']).grid(row=2, column=0, sticky='e', padx=10, pady=5)
        
        order_date_frame = tk.Frame(input_frame, bg=COLORS['light_bg'])
        order_date_frame.grid(row=2, column=1, sticky='w', padx=10, pady=5)
        
        order_day_combo = ttk.Combobox(order_date_frame, values=list(range(1, 32)), width=4, font=FONTS['small'], state='readonly')
        order_day_combo.set(today.day)
        order_day_combo.pack(side='left', padx=(0, 2))
        
        tk.Label(order_date_frame, text="/", font=FONTS['normal'], bg=COLORS['light_bg']).pack(side='left')
        
        order_month_combo = ttk.Combobox(order_date_frame, values=list(range(1, 13)), width=4, font=FONTS['small'], state='readonly')
        order_month_combo.set(today.month)
        order_month_combo.pack(side='left', padx=(2, 2))
        
        tk.Label(order_date_frame, text="/", font=FONTS['normal'], bg=COLORS['light_bg']).pack(side='left')
        
        order_year_combo = ttk.Combobox(order_date_frame, values=list(range(today.year, today.year + 3)), width=6, font=FONTS['small'], state='readonly')
        order_year_combo.set(today.year)
        order_year_combo.pack(side='left', padx=(2, 0))
        
        def update_weekday_from_date(event=None):
            try:
                day = order_day_combo.get()
                month = order_month_combo.get()
                year = order_year_combo.get()
                
                if day and month and year:
                    from datetime import datetime
                    selected_date = datetime(int(year), int(month), int(day))
                    weekday = selected_date.isoweekday()
                    
                    if weekday <= 5:
                        day_combo.set(weekday + 1)
                    elif weekday == 6:
                        day_combo.set(2)
                    else:
                        day_combo.set(2)
            except:
                pass
        
        order_day_combo.bind('<<ComboboxSelected>>', update_weekday_from_date)
        order_month_combo.bind('<<ComboboxSelected>>', update_weekday_from_date)
        order_year_combo.bind('<<ComboboxSelected>>', update_weekday_from_date)
        
        update_weekday_from_date()
        
        tk.Label(input_frame, text="Ghi chú:", font=FONTS['normal'], bg=COLORS['light_bg']).grid(row=2, column=2, sticky='e', padx=10, pady=5)
        notes_entry = tk.Entry(input_frame, width=50, font=FONTS['normal'])
        notes_entry.grid(row=2, column=3, columnspan=3, padx=10, pady=5, sticky='w')
        
        result_frame = tk.Frame(dialog, bg=COLORS['white'])
        result_frame.pack(expand=True, fill='both', padx=20, pady=10)
        
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Order.Treeview", background=COLORS['white'], foreground="black", rowheight=28, fieldbackground=COLORS['white'], font=FONTS['table'])
        style.configure("Order.Treeview.Heading", font=FONTS['table_header'], background=COLORS['darker'], foreground=COLORS['white'], relief='flat')
        style.map("Order.Treeview.Heading", background=[('active', COLORS['darker'])], foreground=[('active', COLORS['white'])])
        style.map('Order.Treeview', background=[('selected', COLORS['primary'])], foreground=[('selected', 'white')])
        
        scrollbar = ttk.Scrollbar(result_frame, orient='vertical')
        scrollbar.pack(side='right', fill='y')
        
        order_tree = ttk.Treeview(
            result_frame,
            columns=('STT', 'Product', 'Qty', 'Portions', 'TotalQty', 'Price', 'Total', 'BeforeVAT'),
            show='headings',
            yscrollcommand=scrollbar.set,
            style="Order.Treeview"
        )
        scrollbar.config(command=order_tree.yview)
        
        order_tree.heading('STT', text='STT', anchor='center')
        order_tree.heading('Product', text='Tên thực phẩm', anchor='w')
        order_tree.heading('Qty', text='Định mức Kg', anchor='center')
        order_tree.heading('Portions', text='Số phần ăn', anchor='center')
        order_tree.heading('TotalQty', text='Đơn vị Kg', anchor='center')
        order_tree.heading('Price', text='Đơn giá', anchor='center')
        order_tree.heading('Total', text='Thành tiền', anchor='center')
        order_tree.heading('BeforeVAT', text='Trước VAT', anchor='center')
        
        order_tree.column('STT', width=50, anchor='center', stretch=False)
        order_tree.column('Product', width=300, anchor='w', stretch=True)
        order_tree.column('Qty', width=120, anchor='center', stretch=False)
        order_tree.column('Portions', width=110, anchor='center', stretch=False)
        order_tree.column('TotalQty', width=110, anchor='center', stretch=False)
        order_tree.column('Price', width=130, anchor='center', stretch=False)
        order_tree.column('Total', width=150, anchor='center', stretch=False)
        order_tree.column('BeforeVAT', width=130, anchor='center', stretch=False)
        
        order_tree.pack(expand=True, fill='both')
        
        order_tree.tag_configure('lunch_item', background='#d4edda', foreground='black')
        order_tree.tag_configure('dinner_item', background='#fff3cd', foreground='black')
        order_tree.tag_configure('total_row', background='#c8e6c9', foreground='black', font=FONTS['table_header'])
        order_tree.tag_configure('lunch_avg', background='#a8d5a8', foreground='black', font=FONTS['normal'])
        order_tree.tag_configure('dinner_avg', background='#9fd4f2', foreground='black', font=FONTS['normal'])
        order_tree.tag_configure('overall_avg', background='#fff59d', foreground='black', font=FONTS['table_header'])
        
        order_items = []
        
        def refresh_order_tree():
            for item in order_tree.get_children():
                order_tree.delete(item)
            
            lunch_total = 0
            dinner_total = 0
            
            for idx, item_data in enumerate(order_items):
                stt = idx + 1
                meal_type = item_data['meal_type']
                product_name = item_data['product_name']
                qty = item_data['qty']
                unit = item_data['unit']
                portions = item_data['portions']
                price = item_data['price']
                
                total_qty = qty * portions
                item_total = qty * portions * price
                before_vat = price / 1.05
                
                if meal_type == 'lunch':
                    lunch_total += item_total
                    tag = 'lunch_item'
                else:
                    dinner_total += item_total
                    tag = 'dinner_item'
                
                order_tree.insert('', 'end', values=(
                    stt,
                    product_name,
                    f"{qty:.2f}",
                    portions,
                    f"{total_qty:.2f}",
                    f"{price:,.0f}",
                    f"{item_total:,.0f}",
                    f"{before_vat:,.0f}"
                ), tags=(tag,))
            
            grand_total = lunch_total + dinner_total
            order_tree.insert('', 'end', values=('', 'TỔNG CỘNG', '', '', '', '', f"{grand_total:,.0f}", ''), tags=('total_row',))
            
            lunch_people_val = lunch_people_entry.get()
            dinner_people_val = dinner_people_entry.get()
            
            try:
                lunch_people_int = int(lunch_people_val) if lunch_people_val else 0
                dinner_people_int = int(dinner_people_val) if dinner_people_val else 0
                
                if lunch_people_int > 0:
                    avg_lunch = lunch_total / lunch_people_int
                    order_tree.insert('', 'end', values=('', 'Đơn giá phần ăn trưa', '', '', '', '', f"{avg_lunch:,.0f}", ''), tags=('lunch_avg',))
                
                if dinner_people_int > 0:
                    avg_dinner = dinner_total / dinner_people_int
                    order_tree.insert('', 'end', values=('', 'Đơn giá phần ăn chiều', '', '', '', '', f"{avg_dinner:,.0f}", ''), tags=('dinner_avg',))
                
                total_people = lunch_people_int + dinner_people_int
                if total_people > 0:
                    overall_avg = grand_total / total_people
                    order_tree.insert('', 'end', values=('', 'TRUNG BÌNH', '', '', '', '', f"{overall_avg:,.0f}", ''), tags=('overall_avg',))
            except ValueError:
                pass
        
        def load_menu_data():
            month = month_combo.get()
            week = week_combo.get()
            day = day_combo.get()
            lunch_people = lunch_people_entry.get()
            dinner_people = dinner_people_entry.get()
            
            if not month or not week or not day:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn tháng, tuần và thứ!")
                return
            
            if not lunch_people or not dinner_people:
                messagebox.showwarning("Cảnh báo", "Vui lòng nhập số người ăn trưa và chiều!")
                return
            
            try:
                lunch_people = int(lunch_people)
                dinner_people = int(dinner_people)
            except ValueError:
                messagebox.showerror("Lỗi", "Số người ăn phải là số nguyên!")
                return
            
            actual_week = week
            if week == '3':
                actual_week = '1'
            elif week == '4':
                actual_week = '2'
            
            menu_name = f"T{month}-{actual_week}-{day}"
            
            menu_items = self.data_manager.get_menu_items(menu_name)
            if not menu_items or (not menu_items.get('lunch') and not menu_items.get('dinner')):
                messagebox.showwarning("Không tìm thấy", f"Không tìm thấy menu '{menu_name}'!")
                return
            
            order_items.clear()
            
            lunch_items = menu_items.get('lunch', [])
            dinner_items = menu_items.get('dinner', [])
            
            if lunch_items:
                for item in lunch_items:
                    order_items.append({
                        'meal_type': 'lunch',
                        'product_name': item['product_name'],
                        'qty': item['qty'],
                        'unit': item['unit'],
                        'portions': lunch_people,
                        'price': item['price']
                    })
            
            if dinner_items:
                for item in dinner_items:
                    order_items.append({
                        'meal_type': 'dinner',
                        'product_name': item['product_name'],
                        'qty': item['qty'],
                        'unit': item['unit'],
                        'portions': dinner_people,
                        'price': item['price']
                    })
            
            refresh_order_tree()
        
        def load_order_data():
            from openpyxl import load_workbook
            import os
            
            day_order = order_day_combo.get()
            month_order = order_month_combo.get()
            year = order_year_combo.get()
            
            if not day_order or not month_order or not year:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn ngày đặt hàng!")
                return
            
            excel_folder = "excel_files"
            file_name = f"DatThucPham{month_order}-{year}.xlsx"
            file_path = os.path.join(excel_folder, file_name)
            
            if not os.path.exists(file_path):
                messagebox.showwarning("Không tìm thấy", f"Không tìm thấy file báo cáo:\n{file_path}")
                return
            
            sheet_name = f"{day_order}-{month_order}-{year}"
            
            try:
                wb = load_workbook(file_path)
                if sheet_name not in wb.sheetnames:
                    messagebox.showwarning("Không tìm thấy", f"Không tìm thấy báo cáo cho ngày {sheet_name}")
                    wb.close()
                    return
                
                ws = wb[sheet_name]
                
                order_items.clear()
                
                lunch_people_val = 0
                dinner_people_val = 0
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[0] or row[1] in ['TỔNG CỘNG', 'Đơn giá phần ăn trưa', 'Đơn giá phần ăn chiều', 'TRUNG BÌNH']:
                        continue
                    
                    product_name = row[1]
                    meal_type = row[2]
                    qty = float(row[3]) if row[3] else 0
                    portions = int(row[4]) if row[4] else 0
                    unit = ws.cell(row=row_idx, column=6).value
                    price = float(row[6]) if row[6] else 0
                    
                    if meal_type == 'Trưa':
                        meal_type_code = 'lunch'
                        lunch_people_val = portions
                    else:
                        meal_type_code = 'dinner'
                        dinner_people_val = portions
                    
                    order_items.append({
                        'meal_type': meal_type_code,
                        'product_name': product_name,
                        'qty': qty,
                        'unit': unit,
                        'portions': portions,
                        'price': price
                    })
                
                if lunch_people_val > 0:
                    lunch_people_entry.delete(0, tk.END)
                    lunch_people_entry.insert(0, str(lunch_people_val))
                
                if dinner_people_val > 0:
                    dinner_people_entry.delete(0, tk.END)
                    dinner_people_entry.insert(0, str(dinner_people_val))
                
                wb.close()
                refresh_order_tree()
                messagebox.showinfo("Thành công", f"Đã tải dữ liệu từ báo cáo ngày {sheet_name}")
                
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể đọc file báo cáo:\n{str(e)}")
        
        def show_add_edit_dialog(edit_index=None):
            edit_dialog = tk.Toplevel(dialog)
            edit_dialog.title("Sửa nguyên liệu" if edit_index is not None else "Thêm nguyên liệu")
            edit_dialog.geometry("500x500")
            edit_dialog.configure(bg=COLORS['white'])
            edit_dialog.transient(dialog)
            edit_dialog.grab_set()
            
            screen_width = edit_dialog.winfo_screenwidth()
            screen_height = edit_dialog.winfo_screenheight()
            x = (screen_width - 500) // 2
            y = (screen_height - 500) // 2
            edit_dialog.geometry(f"500x500+{x}+{y}")
            
            header_frame = tk.Frame(edit_dialog, bg=COLORS['primary'], height=50)
            header_frame.pack(fill='x')
            header_frame.pack_propagate(False)
            
            tk.Label(
                header_frame,
                text="SỬA NGUYÊN LIỆU" if edit_index is not None else "THÊM NGUYÊN LIỆU",
                font=FONTS['header'],
                bg=COLORS['primary'],
                fg=COLORS['white']
            ).pack(pady=10)
            
            form_frame = tk.Frame(edit_dialog, bg=COLORS['white'])
            form_frame.pack(fill='both', expand=True, padx=30, pady=20)
            
            tk.Label(form_frame, text="Loại bữa ăn:", font=FONTS['normal'], bg=COLORS['white']).grid(row=0, column=0, sticky='e', padx=10, pady=10)
            meal_type_combo = ttk.Combobox(form_frame, values=['Trưa', 'Chiều'], width=25, font=FONTS['normal'], state='readonly')
            meal_type_combo.grid(row=0, column=1, padx=10, pady=10, sticky='w')
            
            tk.Label(form_frame, text="Tên thực phẩm:", font=FONTS['normal'], bg=COLORS['white']).grid(row=1, column=0, sticky='e', padx=10, pady=10)
            product_entry = tk.Entry(form_frame, width=28, font=FONTS['normal'])
            product_entry.grid(row=1, column=1, padx=10, pady=10, sticky='w')
            
            tk.Label(form_frame, text="Định mức:", font=FONTS['normal'], bg=COLORS['white']).grid(row=2, column=0, sticky='e', padx=10, pady=10)
            qty_entry = tk.Entry(form_frame, width=28, font=FONTS['normal'])
            qty_entry.grid(row=2, column=1, padx=10, pady=10, sticky='w')
            
            tk.Label(form_frame, text="Đơn vị:", font=FONTS['normal'], bg=COLORS['white']).grid(row=3, column=0, sticky='e', padx=10, pady=10)
            unit_entry = tk.Entry(form_frame, width=28, font=FONTS['normal'])
            unit_entry.grid(row=3, column=1, padx=10, pady=10, sticky='w')
            
            tk.Label(form_frame, text="Số phần ăn:", font=FONTS['normal'], bg=COLORS['white']).grid(row=4, column=0, sticky='e', padx=10, pady=10)
            portions_entry = tk.Entry(form_frame, width=28, font=FONTS['normal'])
            portions_entry.grid(row=4, column=1, padx=10, pady=10, sticky='w')
            
            tk.Label(form_frame, text="Đơn giá:", font=FONTS['normal'], bg=COLORS['white']).grid(row=5, column=0, sticky='e', padx=10, pady=10)
            price_entry = tk.Entry(form_frame, width=28, font=FONTS['normal'])
            price_entry.grid(row=5, column=1, padx=10, pady=10, sticky='w')
            
            if edit_index is not None and 0 <= edit_index < len(order_items):
                item_data = order_items[edit_index]
                meal_type_combo.set('Trưa' if item_data['meal_type'] == 'lunch' else 'Chiều')
                meal_type_combo.config(state='disabled')
                product_entry.insert(0, item_data['product_name'])
                product_entry.config(state='disabled')
                qty_entry.insert(0, str(item_data['qty']))
                unit_entry.insert(0, item_data['unit'])
                unit_entry.config(state='disabled')
                portions_entry.insert(0, str(item_data['portions']))
                portions_entry.config(state='disabled')
                price_entry.insert(0, f"{item_data['price']:,.0f}")
            else:
                meal_type_combo.set('Trưa')
            
            def save_item():
                qty_str = qty_entry.get().strip()
                price_str = price_entry.get().strip()
                
                if not qty_str or not price_str:
                    messagebox.showwarning("Cảnh báo", "Vui lòng điền đầy đủ thông tin!")
                    return
                
                try:
                    qty = float(qty_str)
                    price = float(price_str.replace(',', ''))
                except ValueError:
                    messagebox.showerror("Lỗi", "Định mức và đơn giá phải là số hợp lệ!")
                    return
                
                if edit_index is not None:
                    order_items[edit_index]['qty'] = qty
                    order_items[edit_index]['price'] = price
                else:
                    meal_type_str = meal_type_combo.get()
                    product_name = product_entry.get().strip()
                    unit = unit_entry.get().strip()
                    portions_str = portions_entry.get().strip()
                    
                    if not all([meal_type_str, product_name, unit, portions_str]):
                        messagebox.showwarning("Cảnh báo", "Vui lòng điền đầy đủ thông tin!")
                        return
                    
                    try:
                        portions = int(portions_str)
                    except ValueError:
                        messagebox.showerror("Lỗi", "Số phần ăn phải là số nguyên!")
                        return
                    
                    meal_type = 'lunch' if meal_type_str == 'Trưa' else 'dinner'
                    
                    item_data = {
                        'meal_type': meal_type,
                        'product_name': product_name,
                        'qty': qty,
                        'unit': unit,
                        'portions': portions,
                        'price': price
                    }
                    order_items.append(item_data)
                
                refresh_order_tree()
                edit_dialog.destroy()
            
            button_frame = tk.Frame(edit_dialog, bg=COLORS['white'])
            button_frame.pack(pady=20)
            
            tk.Button(
                button_frame,
                text="💾 Lưu",
                font=FONTS['button'],
                bg=COLORS['success'],
                fg=COLORS['white'],
                width=12,
                height=2,
                command=save_item
            ).pack(side='left', padx=10)
            
            tk.Button(
                button_frame,
                text="✗ Hủy",
                font=FONTS['button'],
                bg=COLORS['secondary'],
                fg=COLORS['white'],
                width=12,
                height=2,
                command=edit_dialog.destroy
            ).pack(side='left', padx=10)
        
        def edit_selected_item():
            selected = order_tree.selection()
            if not selected:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn một nguyên liệu để sửa!")
                return
            
            item_id = selected[0]
            item_values = order_tree.item(item_id, 'values')
            
            if item_values[0] == '' or item_values[1] in ['TỔNG CỘNG', 'Đơn giá phần ăn trưa', 'Đơn giá phần ăn chiều', 'TRUNG BÌNH']:
                messagebox.showwarning("Cảnh báo", "Không thể sửa dòng tổng hợp!")
                return
            
            try:
                stt = int(item_values[0])
                edit_index = stt - 1
                show_add_edit_dialog(edit_index)
            except (ValueError, IndexError):
                messagebox.showerror("Lỗi", "Không thể xác định nguyên liệu!")
        
        def delete_selected_item():
            selected = order_tree.selection()
            if not selected:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn một nguyên liệu để xóa!")
                return
            
            item_id = selected[0]
            item_values = order_tree.item(item_id, 'values')
            
            if item_values[0] == '' or item_values[1] in ['TỔNG CỘNG', 'Đơn giá phần ăn trưa', 'Đơn giá phần ăn chiều', 'TRUNG BÌNH']:
                messagebox.showwarning("Cảnh báo", "Không thể xóa dòng tổng hợp!")
                return
            
            try:
                stt = int(item_values[0])
                delete_index = stt - 1
                
                if messagebox.askyesno("Xác nhận", f"Bạn có chắc muốn xóa '{item_values[1]}'?"):
                    order_items.pop(delete_index)
                    refresh_order_tree()
            except (ValueError, IndexError):
                messagebox.showerror("Lỗi", "Không thể xác định nguyên liệu!")
        
        def on_double_click(event):
            edit_selected_item()
        
        order_tree.bind('<Double-Button-1>', on_double_click)
        

        def export_to_excel():
            if not order_items:
                messagebox.showwarning("Cảnh báo", "Không có dữ liệu để xuất!")
                return
            
            try:
                from openpyxl import Workbook, load_workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                
                day = order_day_combo.get()
                month = order_month_combo.get()
                year = order_year_combo.get()
                date_str = f"{day}-{month}-{year}"
                
                file_path = f"excel_files/DatThucPham{month}-{year}.xlsx"
                
                if os.path.exists(file_path):
                    wb = load_workbook(file_path)
                else:
                    wb = Workbook()
                    if 'Sheet' in wb.sheetnames:
                        del wb['Sheet']
                
                sheet_name = f"{day}-{month}-{year}"
                
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                
                ws = wb.create_sheet(sheet_name)
                
                header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                title_font = Font(bold=True, size=16, color="0066CC")
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                title_text = f"ĐẶT HÀNG NGÀY {day}/{month}/{year}"
                ws.append([title_text])
                ws.merge_cells('A1:H1')
                title_cell = ws['A1']
                title_cell.font = title_font
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[1].height = 30
                
                ws.append([])
                
                headers = ['STT', 'Tên thực phẩm', 'Định mức Kg', 'Số phần ăn', 'Đơn vị Kg', 'Đơn giá', 'Thành tiền', 'Trước VAT']
                ws.append(headers)
                
                for cell in ws[3]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                lunch_total = 0
                dinner_total = 0
                
                lunch_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                dinner_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
                
                for idx, item_data in enumerate(order_items):
                    stt = idx + 1
                    product_name = item_data['product_name']
                    qty = item_data['qty']
                    unit = item_data['unit']
                    portions = item_data['portions']
                    price = item_data['price']
                    meal_type = item_data['meal_type']
                    
                    total_qty = qty * portions
                    item_total = qty * portions * price
                    before_vat = price / 1.05
                    
                    if meal_type == 'lunch':
                        lunch_total += item_total
                        row_fill = lunch_fill
                    else:
                        dinner_total += item_total
                        row_fill = dinner_fill
                    
                    row = [stt, product_name, qty, portions, total_qty, price, item_total, before_vat]
                    ws.append(row)
                    
                    for cell in ws[ws.max_row]:
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = row_fill
                    
                    ws.cell(row=ws.max_row, column=2).alignment = Alignment(horizontal='left', vertical='center')
                    ws.cell(row=ws.max_row, column=3).number_format = '#,##0.00'
                    ws.cell(row=ws.max_row, column=5).number_format = '#,##0.00'
                    ws.cell(row=ws.max_row, column=6).number_format = '#,##0'
                    ws.cell(row=ws.max_row, column=7).number_format = '#,##0'
                    ws.cell(row=ws.max_row, column=8).number_format = '#,##0'
                
                grand_total = lunch_total + dinner_total
                total_row = ['', 'TỔNG CỘNG', '', '', '', '', grand_total, '']
                ws.append(total_row)
                
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                ws.cell(row=ws.max_row, column=7).number_format = '#,##0'
                
                lunch_people_val = lunch_people_entry.get()
                dinner_people_val = dinner_people_entry.get()
                
                if lunch_people_val and int(lunch_people_val) > 0:
                    avg_lunch = lunch_total / int(lunch_people_val)
                    ws.append(['', 'Đơn giá phần ăn trưa', '', '', '', '', avg_lunch, ''])
                    for cell in ws[ws.max_row]:
                        cell.border = border
                    ws.cell(row=ws.max_row, column=7).number_format = '#,##0'
                
                if dinner_people_val and int(dinner_people_val) > 0:
                    avg_dinner = dinner_total / int(dinner_people_val)
                    ws.append(['', 'Đơn giá phần ăn chiều', '', '', '', '', avg_dinner, ''])
                    for cell in ws[ws.max_row]:
                        cell.border = border
                    ws.cell(row=ws.max_row, column=7).number_format = '#,##0'
                
                total_people = int(lunch_people_val or 0) + int(dinner_people_val or 0)
                if total_people > 0:
                    overall_avg = grand_total / total_people
                    ws.append(['', 'TRUNG BÌNH', '', '', '', '', overall_avg, ''])
                    for cell in ws[ws.max_row]:
                        cell.font = Font(bold=True)
                        cell.border = border
                    ws.cell(row=ws.max_row, column=7).number_format = '#,##0'
                
                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 35
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 15
                
                wb.save(file_path)
                messagebox.showinfo("Thành công", f"Đã xuất báo cáo thành công!\n{file_path}")
                
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể xuất báo cáo: {str(e)}")
        
        edit_btn_frame = tk.Frame(dialog, bg=COLORS['white'])
        edit_btn_frame.pack(pady=5)
        
        tk.Button(
            edit_btn_frame,
            text="➕ Thêm",
            font=FONTS['button'],
            bg=COLORS['primary'],
            fg=COLORS['white'],
            width=12,
            height=1,
            command=lambda: show_add_edit_dialog()
        ).pack(side='left', padx=5)
        
        tk.Button(
            edit_btn_frame,
            text="✏️ Sửa",
            font=FONTS['button'],
            bg=COLORS['warning'],
            fg=COLORS['white'],
            width=12,
            height=1,
            command=edit_selected_item
        ).pack(side='left', padx=5)
        
        tk.Button(
            edit_btn_frame,
            text="🗑️ Xóa",
            font=FONTS['button'],
            bg=COLORS['danger'],
            fg=COLORS['white'],
            width=12,
            height=1,
            command=delete_selected_item
        ).pack(side='left', padx=5)
        
        btn_frame = tk.Frame(dialog, bg=COLORS['white'])
        btn_frame.pack(pady=10)
        
        tk.Button(
            btn_frame,
            text="📊 Xem đơn hàng",
            font=FONTS['button'],
            bg=COLORS['success'],
            fg=COLORS['white'],
            width=18,
            height=2,
            command=load_menu_data
        ).pack(side='left', padx=10)
        
        tk.Button(
            btn_frame,
            text="📄 Xuất báo cáo",
            font=FONTS['button'],
            bg='#ff6b6b',
            fg=COLORS['white'],
            width=18,
            height=2,
            command=export_to_excel
        ).pack(side='left', padx=10)
        
        tk.Button(
            btn_frame,
            text="✗ Đóng",
            font=FONTS['button'],
            bg=COLORS['secondary'],
            fg=COLORS['white'],
            width=18,
            height=2,
            command=dialog.destroy
        ).pack(side='left', padx=10)

