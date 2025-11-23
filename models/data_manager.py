from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
from datetime import datetime
from config import EXCEL_FILE

class DataManager:
    def __init__(self):
        self.excel_file = EXCEL_FILE
        self.init_excel()
    
    def init_excel(self):
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            
            ws_products = wb.active
            ws_products.title = "Products"
            ws_products.append(["STT", "TenHang", "DVT", "DonGia"])
            
            ws_fruits = wb.create_sheet("Fruits")
            ws_fruits.append(["STT", "TenHang", "DVT", "DonGia"])
            
            ws_employees = wb.create_sheet("Dat hang thuc phẩm")
            ws_employees.append(["STT", "HO & TEN", "BO PHAN"])
            
            ws_settings = wb.create_sheet("Settings")
            ws_settings.append(["SettingKey", "SettingValue"])
            ws_settings.append(["PageTitle", "CÁC MẶT HÀNG NÔNG SẢN - THỰC PHẨM"])
            
            wb.save(self.excel_file)
        else:
            wb = load_workbook(self.excel_file)
            
            if "Fruits" not in wb.sheetnames:
                ws_fruits = wb.create_sheet("Fruits")
                ws_fruits.append(["STT", "TenHang", "DVT", "DonGia"])
                wb.save(self.excel_file)
            
            if "Dat hang thuc phẩm" not in wb.sheetnames:
                ws_employees = wb.create_sheet("Dat hang thuc phẩm")
                ws_employees.append(["STT", "HO & TEN", "BO PHAN"])
                wb.save(self.excel_file)
            
            wb.close()
    
    def get_next_stt(self, sheet_name="Products"):
        wb = load_workbook(self.excel_file)
        ws = wb[sheet_name]
        max_stt = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and isinstance(row[0], int):
                max_stt = max(max_stt, row[0])
        wb.close()
        return max_stt + 1
    
    def get_page_title(self):
        try:
            wb = load_workbook(self.excel_file)
            ws = wb["Settings"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == "PageTitle":
                    wb.close()
                    return row[1]
            wb.close()
        except:
            pass
        return "CÁC MẶT HÀNG NÔNG SẢN - THỰC PHẨM"
    
    def update_page_title(self, new_title):
        wb = load_workbook(self.excel_file)
        ws = wb["Settings"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == "PageTitle":
                row[1].value = new_title.strip()
                break
        wb.save(self.excel_file)
        wb.close()
    
    def load_data(self, sheet_name="Products"):
        wb = load_workbook(self.excel_file)
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                rows.append(row)
        wb.close()
        rows.sort(key=lambda x: x[0])
        return rows
    
    def add_item(self, sheet_name, name, unit, price):
        stt = self.get_next_stt(sheet_name)
        wb = load_workbook(self.excel_file)
        ws = wb[sheet_name]
        ws.append([stt, name, unit, price])
        wb.save(self.excel_file)
        wb.close()
    
    def update_item(self, sheet_name, stt, name, unit, price):
        wb = load_workbook(self.excel_file)
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == stt:
                row[1].value = name
                row[2].value = unit
                row[3].value = price
                break
        wb.save(self.excel_file)
        wb.close()
    
    def delete_item(self, sheet_name, stt):
        wb = load_workbook(self.excel_file)
        ws = wb[sheet_name]
        row_to_delete = None
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == stt:
                row_to_delete = idx
                break
        if row_to_delete:
            ws.delete_rows(row_to_delete)
            wb.save(self.excel_file)
        wb.close()
    
    def load_employees(self):
        wb = load_workbook(self.excel_file)
        ws = wb["Dat hang thuc phẩm"]
        employees = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                employees.append({
                    'stt': row[0],
                    'name': row[1],
                    'department': row[2]
                })
        wb.close()
        employees.sort(key=lambda x: x['stt'])
        return employees
    
    def get_monthly_meal_file(self, date_obj):
        month = date_obj.strftime("%m")
        year = date_obj.strftime("%Y")
        return f"DAT HANG TU {month}-{year}.xlsx"
    
    def load_meal_registration(self, date_obj):
        meal_file = self.get_monthly_meal_file(date_obj)
        sheet_name = date_obj.strftime("%d-%m-%Y")
        
        if not os.path.exists(meal_file):
            return None
        
        wb = load_workbook(meal_file)
        
        if sheet_name not in wb.sheetnames:
            wb.close()
            return None
        
        ws = wb[sheet_name]
        registrations = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[1]).strip().lower() != 'tổng cộng':
                registrations.append({
                    'stt': row[0],
                    'name': row[1],
                    'department': row[2],
                    'lunch': row[3] if row[3] is not None else 1.0,
                    'dinner': row[4] if row[4] is not None else 1.0,
                    'total': row[5] if row[5] is not None else 2.0,
                    'note': row[6] if row[6] else ''
                })
        wb.close()
        return registrations
    
    def save_meal_registration(self, date_obj, registrations):
        meal_file = self.get_monthly_meal_file(date_obj)
        sheet_name = date_obj.strftime("%d-%m-%Y")
        
        wb = None
        try:
            if os.path.exists(meal_file):
                wb = load_workbook(meal_file)
            else:
                wb = Workbook()
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']
                
                month_str = date_obj.strftime("%m")
                year_str = date_obj.strftime("%Y")
                info_sheet = wb.create_sheet("Thông tin")
                info_sheet.append([f"FILE ĐĂNG KÝ SUẤT ĂN THÁNG {month_str}/{year_str}"])
                info_sheet.merge_cells('A1:D1')
                info_sheet['A1'].font = Font(bold=True, size=16)
                info_sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
                info_sheet.row_dimensions[1].height = 30
                
                info_sheet.append([])
                info_sheet.append(["Ghi chú:", "File này tự động tạo sheet cho mỗi ngày khi đăng ký suất ăn"])
                info_sheet['A3'].font = Font(italic=True)
            
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            
            ws = wb.create_sheet(sheet_name)
            
            title_date = date_obj.strftime("%d/%m/%Y")
            ws.append([f"DANH SÁCH CNV ĐĂNG KÝ SUẤT ĂN NGÀY {title_date}"])
            ws.merge_cells('A1:G1')
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 25
            
            ws.append(["(Thực đơn có thể thay đổi hoặc điều chỉnh theo tình thời điểm thích hợp)"])
            ws.merge_cells('A2:G2')
            ws['A2'].font = Font(italic=True, size=10)
            ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 20
            
            ws.append([])
            
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            headers = ["STT", "HỌ & TÊN", "BỘ PHẬN", "ĂN TRƯA 11h30", "ĂN CHIỀU 16h00", "PHẦN ĂN", "GHI CHÚ"]
            ws.append(headers)
            
            for cell in ws[4]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            ws.row_dimensions[4].height = 35
            
            total_lunch = 0
            total_dinner = 0
            total_meals = 0
            
            for idx, reg in enumerate(registrations, start=5):
                lunch_val = float(reg['lunch'])
                dinner_val = float(reg['dinner'])
                total_val = lunch_val + dinner_val
                
                total_lunch += lunch_val
                total_dinner += dinner_val
                total_meals += total_val
                
                ws.append([
                    reg['stt'],
                    reg['name'],
                    reg['department'],
                    lunch_val,
                    dinner_val,
                    total_val,
                    reg['note']
                ])
                
                for cell in ws[idx]:
                    cell.border = border
                    if cell.column in [1, 4, 5, 6]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
            
            total_row = len(registrations) + 5
            ws.append(['', 'Tổng cộng', '', total_lunch, total_dinner, total_meals, ''])
            
            total_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            total_font = Font(bold=True, size=11)
            
            for cell in ws[total_row]:
                cell.font = total_font
                cell.fill = total_fill
                cell.border = border
                if cell.column in [4, 5, 6]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 16
            ws.column_dimensions['E'].width = 16
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 30
            
            wb.save(meal_file)
            return meal_file
        except Exception as e:
            raise Exception(f"Lỗi khi lưu file: {str(e)}")
        finally:
            if wb:
                wb.close()
    
    def load_meal_list(self):
        try:
            wb = load_workbook(self.excel_file)
            sheet_name = "Meal_20251120"
            
            if sheet_name not in wb.sheetnames:
                wb.close()
                return []
            
            ws = wb[sheet_name]
            meal_list = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    meal_list.append({
                        'stt': row[0],
                        'name': row[1] if len(row) > 1 else '',
                        'department': row[2] if len(row) > 2 else '',
                        'lunch': row[3] if len(row) > 3 and row[3] is not None else 1.0,
                        'dinner': row[4] if len(row) > 4 and row[4] is not None else 1.0,
                        'total': row[5] if len(row) > 5 and row[5] is not None else 2.0,
                        'note': row[6] if len(row) > 6 and row[6] else ''
                    })
            wb.close()
            return meal_list
        except Exception as e:
            raise Exception(f"Lỗi khi đọc dữ liệu từ sheet Meal_20251120: {str(e)}")
    
    def save_meal_list(self, meal_list):
        try:
            wb = load_workbook(self.excel_file)
            sheet_name = "Meal_20251120"
            
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            
            ws = wb.create_sheet(sheet_name)
            
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            headers = ["STT", "HỌ & TÊN", "BỘ PHẬN", "ĂN TRƯA\n11h30", "ĂN CHIỀU\n16h00", "PHẦN ĂN", "GHI CHÚ"]
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            ws.row_dimensions[1].height = 35
            
            for idx, meal in enumerate(meal_list, start=2):
                lunch_val = float(meal['lunch'])
                dinner_val = float(meal['dinner'])
                total_val = lunch_val + dinner_val
                
                ws.append([
                    meal['stt'],
                    meal['name'],
                    meal['department'],
                    lunch_val,
                    dinner_val,
                    total_val,
                    meal['note']
                ])
                
                for cell in ws[idx]:
                    cell.border = border
                    if cell.column in [1, 4, 5, 6]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
            
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 16
            ws.column_dimensions['E'].width = 16
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 30
            
            wb.save(self.excel_file)
            wb.close()
        except Exception as e:
            raise Exception(f"Lỗi khi lưu dữ liệu vào sheet Meal_20251120: {str(e)}")
    
    def save_menu_summary(self, date_obj, morning_items, evening_items, total_lunch, total_dinner):
        meal_file = self.get_monthly_meal_file(date_obj)
        sheet_name = date_obj.strftime("%d-%m-%Y")
        
        if not os.path.exists(meal_file):
            raise Exception(f"File {meal_file} không tồn tại. Vui lòng lưu đăng ký suất ăn trước.")
        
        wb = load_workbook(meal_file)
        
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise Exception(f"Sheet {sheet_name} không tồn tại. Vui lòng lưu đăng ký suất ăn trước.")
        
        ws = wb[sheet_name]
        
        last_row = ws.max_row
        start_row = last_row + 3
        
        title_date = date_obj.strftime("%d/%m/%Y")
        ws.merge_cells(f'A{start_row}:H{start_row}')
        ws[f'A{start_row}'] = f"ĐẶT HÀNG THỰC PHẨM NGÀY {title_date}"
        ws[f'A{start_row}'].font = Font(bold=True, size=14)
        ws[f'A{start_row}'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[start_row].height = 25
        
        start_row += 1
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        headers = ["TT", "Tên thực phẩm", "Định mức Kg", "Số phần ăn", "Đơn vị Kg", "Đơn giá/Kg", "Thành tiền", "Trước VAT"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        ws.row_dimensions[start_row].height = 35
        current_row = start_row + 1
        
        morning_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        evening_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        
        total_morning_price = 0
        total_evening_price = 0
        
        stt = 1
        for item in morning_items:
            ws.cell(row=current_row, column=1, value=stt)
            ws.cell(row=current_row, column=2, value=item['name'])
            ws.cell(row=current_row, column=3, value=item['qty'])
            ws.cell(row=current_row, column=3).number_format = '0.00'
            ws.cell(row=current_row, column=4, value=int(total_lunch))
            ws.cell(row=current_row, column=5, value=item['unit'])
            ws.cell(row=current_row, column=5).number_format = '0.00'
            ws.cell(row=current_row, column=6, value=item['price'])
            ws.cell(row=current_row, column=6).number_format = '#,##0'
            ws.cell(row=current_row, column=7, value=item['total'])
            ws.cell(row=current_row, column=7).number_format = '#,##0'
            ws.cell(row=current_row, column=8, value=item['vat'])
            ws.cell(row=current_row, column=8).number_format = '#,##0'
            
            for col_idx in range(1, 9):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = border
                cell.fill = morning_fill
                if col_idx in [1, 3, 4, 5, 6, 7, 8]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            total_morning_price += item['total']
            current_row += 1
            stt += 1
        
        for item in evening_items:
            ws.cell(row=current_row, column=1, value=stt)
            ws.cell(row=current_row, column=2, value=item['name'])
            ws.cell(row=current_row, column=3, value=item['qty'])
            ws.cell(row=current_row, column=3).number_format = '0.00'
            ws.cell(row=current_row, column=4, value=int(total_dinner))
            ws.cell(row=current_row, column=5, value=item['unit'])
            ws.cell(row=current_row, column=5).number_format = '0.00'
            ws.cell(row=current_row, column=6, value=item['price'])
            ws.cell(row=current_row, column=6).number_format = '#,##0'
            ws.cell(row=current_row, column=7, value=item['total'])
            ws.cell(row=current_row, column=7).number_format = '#,##0'
            ws.cell(row=current_row, column=8, value=item['vat'])
            ws.cell(row=current_row, column=8).number_format = '#,##0'
            
            for col_idx in range(1, 9):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = border
                cell.fill = evening_fill
                if col_idx in [1, 3, 4, 5, 6, 7, 8]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            total_evening_price += item['total']
            current_row += 1
            stt += 1
        
        total_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        total_font = Font(bold=True, size=11)
        
        total_price = total_morning_price + total_evening_price
        ws.cell(row=current_row, column=2, value="Tổng cộng")
        ws.cell(row=current_row, column=7, value=total_price)
        ws.cell(row=current_row, column=7).number_format = '#,##0'
        
        for col_idx in range(1, 9):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
        
        morning_avg_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        morning_avg = total_morning_price / total_lunch if total_lunch > 0 else 0
        ws.cell(row=current_row, column=2, value="Đơn giá phần ăn trưa")
        ws.cell(row=current_row, column=7, value=morning_avg)
        ws.cell(row=current_row, column=7).number_format = '#,##0'
        
        for col_idx in range(1, 9):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = total_font
            cell.fill = morning_avg_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
        
        evening_avg_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        evening_avg = total_evening_price / total_dinner if total_dinner > 0 else 0
        ws.cell(row=current_row, column=2, value="Đơn giá phần ăn chiều")
        ws.cell(row=current_row, column=7, value=evening_avg)
        ws.cell(row=current_row, column=7).number_format = '#,##0'
        
        for col_idx in range(1, 9):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = total_font
            cell.fill = evening_avg_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
        
        average_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        total_meals = total_lunch + total_dinner
        overall_avg = total_price / total_meals if total_meals > 0 else 0
        ws.cell(row=current_row, column=2, value="TRUNG BÌNH")
        ws.cell(row=current_row, column=7, value=overall_avg)
        ws.cell(row=current_row, column=7).number_format = '#,##0'
        
        for col_idx in range(1, 9):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = total_font
            cell.fill = average_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        
        wb.save(meal_file)
        wb.close()
    
    def get_monthly_statistics_data(self, month, year):
        import calendar
        from datetime import date
        
        month_str = str(month).zfill(2)
        source_file = f"DAT HANG TU {month_str}-{year}.xlsx"
        
        if not os.path.exists(source_file):
            raise Exception(f"Không tìm thấy file {source_file}")
        
        num_days = calendar.monthrange(year, month)[1]
        wb_source = load_workbook(source_file)
        days_of_week = ["Hai", "Ba", "Tư", "Năm", "Sáu", "Bảy", "CN"]
        
        data_rows = []
        
        for day in range(1, num_days + 1):
            day_str = str(day).zfill(2)
            sheet_date_name = f"{day_str}-{month_str}-{year}"
            
            current_date = date(year, month, day)
            weekday = current_date.weekday()
            day_name = days_of_week[weekday]
            
            total_meals = 0
            total_cost = 0
            
            if sheet_date_name in wb_source.sheetnames:
                ws_day = wb_source[sheet_date_name]
                
                for row in ws_day.iter_rows(min_row=2, values_only=True):
                    if len(row) > 1 and row[1] and str(row[1]).strip().lower() == 'tổng cộng':
                        if len(row) > 3 and row[3] is not None:
                            total_meals += float(row[3])
                        if len(row) > 4 and row[4] is not None:
                            total_meals += float(row[4])
                        break
                
                for row_idx, row in enumerate(ws_day.iter_rows(min_row=1, values_only=True), 1):
                    if len(row) > 0 and row[0] and 'ĐẶT HÀNG THỰC PHẨM' in str(row[0]).upper():
                        for cost_row in ws_day.iter_rows(min_row=row_idx + 1, values_only=True):
                            if len(cost_row) > 1 and cost_row[1] and str(cost_row[1]).strip().lower() == 'tổng cộng':
                                if len(cost_row) > 6 and cost_row[6] is not None:
                                    total_cost = float(cost_row[6])
                                break
                        break
            
            avg_price = total_cost / total_meals if total_meals > 0 else 0
            
            data_rows.append({
                'day_name': day_name,
                'day': day,
                'total_meals': total_meals,
                'total_cost': total_cost,
                'avg_price': avg_price,
                'weekday': weekday,
                'note': 'Nghỉ lễ' if total_meals == 0 else ''
            })
        
        wb_source.close()
        
        week_avg_prices = []
        for idx, day_data in enumerate(data_rows):
            avg_price = day_data['avg_price']
            weekday = day_data['weekday']
            
            if avg_price > 0:
                week_avg_prices.append(avg_price)
            
            if weekday == 6 or idx == len(data_rows) - 1:
                if week_avg_prices:
                    week_avg = sum(week_avg_prices) / len(week_avg_prices)
                    
                    for back_idx in range(len(week_avg_prices)):
                        data_rows[idx - back_idx]['week_avg'] = week_avg
                    
                    week_avg_prices = []
        
        return data_rows
    
    def generate_monthly_statistics(self, month, year):
        import calendar
        from datetime import date
        
        month_str = str(month).zfill(2)
        source_file = f"DAT HANG TU {month_str}-{year}.xlsx"
        
        if not os.path.exists(source_file):
            raise Exception(f"Không tìm thấy file {source_file}")
        
        stats_file = f"Tong hop chi phi trong thang nam {year}.xlsx"
        
        if os.path.exists(stats_file):
            wb_stats = load_workbook(stats_file)
        else:
            wb_stats = Workbook()
            if 'Sheet' in wb_stats.sheetnames:
                del wb_stats['Sheet']
        
        sheet_name = f"Thang {month_str}"
        if sheet_name in wb_stats.sheetnames:
            del wb_stats[sheet_name]
        
        ws_stats = wb_stats.create_sheet(sheet_name)
        
        ws_stats.merge_cells('A1:G1')
        ws_stats['A1'] = f"THEO DÕI CHI PHÍ SUẤT ĂN THÁNG {month_str}/{year}"
        ws_stats['A1'].font = Font(bold=True, size=14)
        ws_stats['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws_stats.row_dimensions[1].height = 25
        
        ws_stats.append([])
        ws_stats.append([])
        
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = ["Thứ", "Ngày", "Số phần ăn\nTrưa + Chiều", "Chi phí\nthực phẩm", "Đơn giá\nmỗi suất ăn", "Trung bình suất\năn của tuần", "Ghi chú"]
        ws_stats.append(headers)
        
        for cell in ws_stats[4]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        ws_stats.row_dimensions[4].height = 40
        
        num_days = calendar.monthrange(year, month)[1]
        
        wb_source = load_workbook(source_file)
        
        days_of_week = ["Hai", "Ba", "Tư", "Năm", "Sáu", "Bảy", "CN"]
        
        data_rows = []
        
        for day in range(1, num_days + 1):
            day_str = str(day).zfill(2)
            sheet_date_name = f"{day_str}-{month_str}-{year}"
            
            current_date = date(year, month, day)
            weekday = current_date.weekday()
            day_name = days_of_week[weekday]
            
            total_meals = 0
            total_cost = 0
            
            if sheet_date_name in wb_source.sheetnames:
                ws_day = wb_source[sheet_date_name]
                
                for row in ws_day.iter_rows(min_row=2, values_only=True):
                    if len(row) > 1 and row[1] and str(row[1]).strip().lower() == 'tổng cộng':
                        if len(row) > 3 and row[3] is not None:
                            total_meals += float(row[3])
                        if len(row) > 4 and row[4] is not None:
                            total_meals += float(row[4])
                        break
                
                has_menu_data = False
                for row_idx, row in enumerate(ws_day.iter_rows(min_row=1, values_only=True), 1):
                    if len(row) > 0 and row[0] and 'ĐẶT HÀNG THỰC PHẨM' in str(row[0]).upper():
                        has_menu_data = True
                        
                        for cost_row in ws_day.iter_rows(min_row=row_idx + 1, values_only=True):
                            if len(cost_row) > 1 and cost_row[1] and str(cost_row[1]).strip().lower() == 'tổng cộng':
                                if len(cost_row) > 6 and cost_row[6] is not None:
                                    total_cost = float(cost_row[6])
                                break
                        break
            
            avg_price = total_cost / total_meals if total_meals > 0 else 0
            
            data_rows.append({
                'day_name': day_name,
                'day': day,
                'total_meals': total_meals,
                'total_cost': total_cost,
                'avg_price': avg_price,
                'weekday': weekday
            })
        
        wb_source.close()
        
        current_row = 5
        week_avg_prices = []
        
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        
        for idx, day_data in enumerate(data_rows):
            day = day_data['day']
            day_name = day_data['day_name']
            total_meals = day_data['total_meals']
            total_cost = day_data['total_cost']
            avg_price = day_data['avg_price']
            weekday = day_data['weekday']
            
            if total_meals == 0:
                ws_stats.append([
                    day_name,
                    day,
                    0,
                    '-',
                    '',
                    '',
                    'Nghỉ lễ' if weekday == 6 else 'Nghỉ lễ'
                ])
                
                for col_idx in range(1, 8):
                    cell = ws_stats.cell(row=current_row, column=col_idx)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if col_idx == 7:
                        cell.font = Font(italic=True, color="FF0000")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    if weekday in [5, 6]:
                        cell.fill = yellow_fill
            else:
                if avg_price > 0:
                    week_avg_prices.append(avg_price)
                
                ws_stats.append([
                    day_name,
                    day,
                    total_meals,
                    total_cost,
                    avg_price if avg_price > 0 else '',
                    '',
                    ''
                ])
                
                for col_idx in range(1, 8):
                    cell = ws_stats.cell(row=current_row, column=col_idx)
                    cell.border = border
                    
                    if col_idx in [1, 2]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_idx in [3, 4, 5, 6]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        if col_idx == 4:
                            cell.number_format = '#,##0'
                        elif col_idx == 5:
                            cell.number_format = '#,##0'
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                    if weekday in [5, 6]:
                        cell.fill = yellow_fill
            
            if weekday == 6 or idx == len(data_rows) - 1:
                if week_avg_prices:
                    week_avg = sum(week_avg_prices) / len(week_avg_prices)
                    
                    for back_idx in range(len(week_avg_prices)):
                        week_row = current_row - back_idx
                        ws_stats.cell(row=week_row, column=6, value=week_avg)
                        ws_stats.cell(row=week_row, column=6).number_format = '#,##0'
                        ws_stats.cell(row=week_row, column=6).alignment = Alignment(horizontal="center", vertical="center")
                    
                    week_avg_prices = []
            
            current_row += 1
        
        ws_stats.column_dimensions['A'].width = 10
        ws_stats.column_dimensions['B'].width = 10
        ws_stats.column_dimensions['C'].width = 18
        ws_stats.column_dimensions['D'].width = 18
        ws_stats.column_dimensions['E'].width = 18
        ws_stats.column_dimensions['F'].width = 18
        ws_stats.column_dimensions['G'].width = 25
        
        wb_stats.save(stats_file)
        wb_stats.close()
        
        return f"Đã tạo thống kê tháng {month}/{year}\nFile: {stats_file}"
